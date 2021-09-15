import datetime
import io
from typing import TYPE_CHECKING, Union

import pytest
from django.utils.timezone import make_aware
from openpyxl import load_workbook

from web.domains.case._import.derogations.models import DerogationsApplication
from web.domains.case._import.fa_dfl.models import DFLApplication
from web.domains.case._import.fa_oil.models import OpenIndividualLicenceApplication
from web.domains.case._import.fa_sil.models import SILApplication
from web.domains.case._import.ironsteel.models import IronSteelApplication
from web.domains.case._import.models import ImportApplication, ImportApplicationType
from web.domains.case._import.opt.models import (
    CP_CATEGORIES,
    OutwardProcessingTradeApplication,
)
from web.domains.case._import.sanctions.models import (
    SanctionsAndAdhocApplication,
    SanctionsAndAdhocApplicationGoods,
)
from web.domains.case._import.sps.models import PriorSurveillanceApplication
from web.domains.case._import.textiles.models import TextilesApplication
from web.domains.case._import.wood.models import WoodQuotaApplication
from web.domains.case.export.models import (
    CertificateOfFreeSaleApplication,
    CertificateOfGoodManufacturingPracticeApplication,
    CertificateOfManufactureApplication,
    ExportApplicationType,
)
from web.domains.case.models import ApplicationBase
from web.domains.case.utils import get_application_current_task
from web.domains.commodity.models import Commodity, CommodityGroup, CommodityType
from web.domains.country.models import Country
from web.flow.models import Task
from web.models.shared import FirearmCommodity, YesNoChoices
from web.types import AuthenticatedHttpRequest, ICMSMiddlewareContext
from web.utils.search import (
    CommodityDetails,
    ResultRow,
    SearchTerms,
    get_search_results_spreadsheet,
    search_applications,
)

if TYPE_CHECKING:
    from web.models import User
    from web.domains.importer.models import Importer
    from web.domains.exporter.models import Exporter

from typing import NamedTuple


class FixtureData(NamedTuple):
    importer: "Importer"
    agent_importer: "Importer"
    importer_user: "User"
    request: AuthenticatedHttpRequest


@pytest.fixture
def test_data(db, importer, agent_importer, test_import_user, request):
    request.user = test_import_user
    request.icms = ICMSMiddlewareContext()

    return FixtureData(importer, agent_importer, test_import_user, request)


class ExportFixtureData(NamedTuple):
    exporter: "Exporter"
    agent_exporter: "Exporter"
    exporter_user: "User"
    request: AuthenticatedHttpRequest


@pytest.fixture
def export_fixture_data(db, exporter, agent_exporter, test_export_user, request):
    request.user = test_export_user
    request.icms = ICMSMiddlewareContext()

    return ExportFixtureData(exporter, agent_exporter, test_export_user, request)


def test_filter_by_application_type(test_data: FixtureData):
    _create_wood_application("Wood ref 1", test_data)
    _create_wood_application("Wood ref 2", test_data)
    _create_derogation_application("Derogation ref 1", test_data)
    _create_derogation_application("Derogation ref 2", test_data)

    terms = SearchTerms(case_type="import", app_type=ImportApplicationType.Types.WOOD_QUOTA)  # type: ignore[arg-type]
    results = search_applications(terms)

    assert results.total_rows == 2

    check_application_references(results.records, "Wood ref 2", "Wood ref 1")


def test_filter_wood(test_data: FixtureData):
    """Do several tests related to searching for wood queries.

    All tests run in a single test for speed.
    The created applications are therefore reused in several tests.
    """

    _test_fetch_all(test_data)

    _test_search_by_case_reference(test_data)

    _test_search_by_status(
        ImportApplicationType.Types.WOOD_QUOTA,
        ImportApplication.Statuses.SUBMITTED,
        expected=["Wood ref 3", "Wood ref 2", "Wood ref 1"],
    )

    _test_search_by_response_decision()

    _test_search_by_importer_or_agent_name(test_data)

    _test_search_by_submitted_datetime(test_data)

    _test_search_by_licence_date()


def test_order_and_limit_works(test_data: FixtureData):
    """Create many applications and ensure only the latest n submitted are returned"""

    applications = (
        "wood app 1",
        "wood app 2",
        "derogation app 1",
        "derogation app 2",
        "wood app 3",
        "wood app 4",
        "wood app 5",
        "derogation app 3",
        "derogation app 4",
        "derogation app 5",
    )

    for app_ref in applications:
        if app_ref.startswith("wood"):
            _create_wood_application(app_ref, test_data)

        elif app_ref.startswith("derogation"):
            _create_derogation_application(app_ref, test_data)

        else:
            raise Exception(f"failed to create: {app_ref}")

    terms = SearchTerms(case_type="import")
    search_data = search_applications(terms, limit=5)

    assert search_data.total_rows == 10

    assert len(search_data.records) == 5

    check_application_references(
        search_data.records,
        "derogation app 5",
        "derogation app 4",
        "derogation app 3",
        "wood app 5",
        "wood app 4",
    )


def test_derogation_commodity_details_correct(test_data: FixtureData):
    app = _create_derogation_application("derogation app 1", test_data)

    search_terms = SearchTerms(case_type="import", app_type=ImportApplicationType.Types.DEROGATION)  # type: ignore[arg-type]
    results = search_applications(search_terms)

    assert results.total_rows == 1
    check_application_references(results.records, "derogation app 1")
    check_commodity_details(
        results.records[0].commodity_details,
        expected_origin_country="Tanzania",
        expected_consignment_country="Algeria",
        expected_shipping_year=app.submit_datetime.year,
        expected_commodity_codes=["code112233"],
    )


def test_fa_dfl_commodity_details_correct(test_data: FixtureData):
    _create_fa_dfl_application("fa-dfl app 1", test_data)

    search_terms = SearchTerms(
        case_type="import",
        app_type=ImportApplicationType.Types.FIREARMS,  # type: ignore[arg-type]
        app_sub_type=ImportApplicationType.SubTypes.DFL,  # type: ignore[arg-type]
    )
    results = search_applications(search_terms)

    assert results.total_rows == 1
    check_application_references(results.records, "fa-dfl app 1")
    check_commodity_details(
        results.records[0].commodity_details,
        expected_origin_country="the Czech Republic",
        expected_consignment_country="the Slovak Republic",
        expected_goods_category=FirearmCommodity.EX_CHAPTER_97.label,  # type: ignore[attr-defined]
    )


def test_fa_oil_commodity_details_correct(test_data: FixtureData):
    _create_fa_oil_application("fa-oil app 1", test_data)

    search_terms = SearchTerms(
        case_type="import",
        app_type=ImportApplicationType.Types.FIREARMS,  # type: ignore[arg-type]
        app_sub_type=ImportApplicationType.SubTypes.OIL,  # type: ignore[arg-type]
    )
    results = search_applications(search_terms)

    assert results.total_rows == 1
    check_application_references(results.records, "fa-oil app 1")
    check_commodity_details(
        results.records[0].commodity_details,
        expected_origin_country="Any Country",
        expected_consignment_country="Any Country",
        expected_goods_category=FirearmCommodity.EX_CHAPTER_93.label,  # type: ignore[attr-defined]
    )


def test_fa_sil_commodity_details_correct(test_data: FixtureData):
    _create_fa_sil_application("fa-sil app 1", test_data)

    search_terms = SearchTerms(
        case_type="import",
        app_type=ImportApplicationType.Types.FIREARMS,  # type: ignore[arg-type]
        app_sub_type=ImportApplicationType.SubTypes.SIL,  # type: ignore[arg-type]
    )
    results = search_applications(search_terms)

    assert results.total_rows == 1
    check_application_references(results.records, "fa-sil app 1")
    check_commodity_details(
        results.records[0].commodity_details,
        expected_origin_country="Argentina",
        expected_consignment_country="Azerbaijan",
        expected_goods_category=FirearmCommodity.EX_CHAPTER_97.label,  # type: ignore[attr-defined]
    )


def test_ironsteel_commodity_details_correct(test_data: FixtureData):
    _create_ironsteel_application("ironsteel app 1", test_data)

    search_terms = SearchTerms(case_type="import", app_type=ImportApplicationType.Types.IRON_STEEL)  # type: ignore[arg-type]
    results = search_applications(search_terms)

    assert results.total_rows == 1
    check_application_references(results.records, "ironsteel app 1")
    check_commodity_details(
        results.records[0].commodity_details,
        expected_origin_country="Kazakhstan",
        expected_consignment_country="Bahamas",
        expected_shipping_year=2021,
        expected_goods_category="SA1",
        expected_commodity_codes=["7208370010"],
    )


def test_opt_commodity_details_correct(test_data: FixtureData):
    app = _create_opt_application("opt app 1", test_data)

    search_terms = SearchTerms(case_type="import", app_type=ImportApplicationType.Types.OPT)  # type: ignore[arg-type]
    results = search_applications(search_terms)

    assert results.total_rows == 1
    check_application_references(results.records, "opt app 1")
    check_commodity_details(
        results.records[0].commodity_details,
        expected_origin_country="Uruguay",
        expected_consignment_country="USA",
        expected_shipping_year=app.submit_datetime.year,
        expected_goods_category=CP_CATEGORIES[0],
        expected_commodity_codes=[
            "5006009000",
            "5007206190",
            "5112301000",
            "6205200010",
            "6205908010",
        ],
    )


def test_sanctionadhoc_commodity_details_correct(test_data: FixtureData):
    app = _create_sanctionadhoc_application("sanctionsadhoc app 1", test_data)

    search_terms = SearchTerms(
        case_type="import", app_type=ImportApplicationType.Types.SANCTION_ADHOC  # type: ignore[arg-type]
    )
    results = search_applications(search_terms)

    assert results.total_rows == 1
    check_application_references(results.records, "sanctionsadhoc app 1")
    check_commodity_details(
        results.records[0].commodity_details,
        expected_origin_country="Iran",
        expected_consignment_country="Algeria",
        expected_shipping_year=app.submit_datetime.year,
        expected_commodity_codes=["2801000010", "2850002070"],
    )


def test_sps_commodity_details_correct(test_data: FixtureData):
    app = _create_sps_application("sps app 1", test_data)

    search_terms = SearchTerms(case_type="import", app_type=ImportApplicationType.Types.SPS)  # type: ignore[arg-type]
    results = search_applications(search_terms)

    assert results.total_rows == 1
    check_application_references(results.records, "sps app 1")
    check_commodity_details(
        results.records[0].commodity_details,
        expected_origin_country="Azerbaijan",
        expected_consignment_country="Jordan",
        expected_shipping_year=app.submit_datetime.year,
        expected_commodity_codes=["7208539000"],
    )


def test_sps_commodity_details_correct_multiple(test_data: FixtureData):
    app_1 = _create_sps_application(
        "sps app 1",
        test_data,
        origin_country="Afghanistan",
        consignment_country="Armenia",
        commodity_code="111111",
    )
    app_2 = _create_sps_application(
        "sps app 2",
        test_data,
        origin_country="Albania",
        consignment_country="Australia",
        commodity_code="222222",
    )
    app_3 = _create_sps_application(
        "sps app 3",
        test_data,
        origin_country="Algeria",
        consignment_country="Azerbaijan",
        commodity_code="333333",
    )
    app_4 = _create_sps_application(
        "sps app 4",
        test_data,
        origin_country="Angola",
        consignment_country="Bahamas",
        commodity_code="444444",
    )
    app_5 = _create_sps_application(
        "sps app 5",
        test_data,
        origin_country="Argentina",
        consignment_country="Bahrain",
        commodity_code="555555",
    )

    search_terms = SearchTerms(case_type="import", app_type=ImportApplicationType.Types.SPS)  # type: ignore[arg-type]
    results = search_applications(search_terms)

    assert results.total_rows == 5
    check_application_references(
        results.records, "sps app 5", "sps app 4", "sps app 3", "sps app 2", "sps app 1"
    )

    check_commodity_details(
        results.records[4].commodity_details,
        expected_origin_country="Afghanistan",
        expected_consignment_country="Armenia",
        expected_shipping_year=app_1.submit_datetime.year,
        expected_commodity_codes=["111111"],
    )

    check_commodity_details(
        results.records[3].commodity_details,
        expected_origin_country="Albania",
        expected_consignment_country="Australia",
        expected_shipping_year=app_2.submit_datetime.year,
        expected_commodity_codes=["222222"],
    )

    check_commodity_details(
        results.records[2].commodity_details,
        expected_origin_country="Algeria",
        expected_consignment_country="Azerbaijan",
        expected_shipping_year=app_3.submit_datetime.year,
        expected_commodity_codes=["333333"],
    )

    check_commodity_details(
        results.records[1].commodity_details,
        expected_origin_country="Angola",
        expected_consignment_country="Bahamas",
        expected_shipping_year=app_4.submit_datetime.year,
        expected_commodity_codes=["444444"],
    )

    check_commodity_details(
        results.records[0].commodity_details,
        expected_origin_country="Argentina",
        expected_consignment_country="Bahrain",
        expected_shipping_year=app_5.submit_datetime.year,
        expected_commodity_codes=["555555"],
    )


def test_textiles_commodity_details_correct(test_data: FixtureData):
    _create_textiles_application("textiles app 1", test_data)

    search_terms = SearchTerms(case_type="import", app_type=ImportApplicationType.Types.TEXTILES)  # type: ignore[arg-type]
    results = search_applications(search_terms)

    assert results.total_rows == 1
    check_application_references(results.records, "textiles app 1")
    check_commodity_details(
        results.records[0].commodity_details,
        expected_origin_country="Belarus",
        expected_consignment_country="Argentina",
        expected_shipping_year=2024,
        expected_goods_category="22",
        expected_commodity_codes=["5509620000"],
    )


def test_wood_commodity_details_correct(test_data: FixtureData):
    _create_wood_application(
        "Wood ref 1", test_data, shipping_year=2030, commodity_code="code654321"
    )

    search_terms = SearchTerms(case_type="import", app_type=ImportApplicationType.Types.WOOD_QUOTA)  # type: ignore[arg-type]
    results = search_applications(search_terms)

    assert results.total_rows == 1

    check_application_references(results.records, "Wood ref 1")

    wood_app = results.records[0]

    assert wood_app.commodity_details.origin_country == "None"
    assert wood_app.commodity_details.shipping_year == 2030
    assert wood_app.commodity_details.commodity_codes == ["code654321"]


def _test_fetch_all(test_data: FixtureData):
    _create_wood_application("Wood in progress", test_data, submit=False)
    _create_wood_application("Wood ref 1", test_data)
    _create_wood_application("Wood ref 2", test_data)

    search_terms = SearchTerms(case_type="import", app_type=ImportApplicationType.Types.WOOD_QUOTA)  # type: ignore[arg-type]
    results = search_applications(search_terms)

    assert results.total_rows == 2

    check_application_references(results.records, "Wood ref 2", "Wood ref 1")


def _test_search_by_case_reference(test_data: FixtureData):
    """Test submitting an application and searching for it by the case reference"""

    application = _create_wood_application("Wood ref 3", test_data)
    case_reference = application.get_reference()

    assert case_reference != "Not Assigned"

    search_terms = SearchTerms(
        case_type="import", app_type=ImportApplicationType.Types.WOOD_QUOTA, case_ref=case_reference  # type: ignore[arg-type]
    )
    results = search_applications(search_terms)

    assert results.total_rows == 1
    check_application_references(results.records, "Wood ref 3")


def _test_search_by_status(app_type: str, case_status: str, expected: list[str]):
    """Search by status using the records we have already created"""

    search_terms = SearchTerms(case_type="import", app_type=app_type, case_status=case_status)
    results = search_applications(search_terms)

    assert results.total_rows == len(expected)
    check_application_references(results.records, *expected)


def _test_search_by_response_decision():
    submitted_application = WoodQuotaApplication.objects.get(applicant_reference="Wood ref 3")
    submitted_application.decision = ApplicationBase.APPROVE
    submitted_application.save()

    search_terms = SearchTerms(
        case_type="import",
        app_type=ImportApplicationType.Types.WOOD_QUOTA,  # type: ignore[arg-type]
        response_decision=ApplicationBase.REFUSE,
    )
    results = search_applications(search_terms)

    assert results.total_rows == 0

    search_terms = SearchTerms(
        case_type="import",
        app_type=ImportApplicationType.Types.WOOD_QUOTA,  # type: ignore[arg-type]
        response_decision=ApplicationBase.APPROVE,
    )
    results = search_applications(search_terms)

    assert results.total_rows == 1

    check_application_references(results.records, "Wood ref 3")


def _test_search_by_importer_or_agent_name(test_data: FixtureData):
    search_terms = SearchTerms(
        case_type="import",
        app_type=ImportApplicationType.Types.WOOD_QUOTA,  # type: ignore[arg-type]
        importer_agent_name="Not valid",
    )
    results = search_applications(search_terms)

    assert results.total_rows == 0

    search_terms = SearchTerms(
        case_type="import",
        app_type=ImportApplicationType.Types.WOOD_QUOTA,  # type: ignore[arg-type]
        importer_agent_name=test_data.importer.name,
    )
    results = search_applications(search_terms)

    assert results.total_rows == 3

    check_application_references(results.records, "Wood ref 3", "Wood ref 2", "Wood ref 1")

    # Set an agent on the first application and check we can search for that.
    application = WoodQuotaApplication.objects.get(applicant_reference="Wood ref 1")
    application.agent = test_data.agent_importer
    application.save()

    search_terms = SearchTerms(
        case_type="import",
        app_type=ImportApplicationType.Types.WOOD_QUOTA,  # type: ignore[arg-type]
        importer_agent_name=test_data.agent_importer.name,
    )
    results = search_applications(search_terms)

    assert results.total_rows == 1

    check_application_references(results.records, "Wood ref 1")


def _test_search_by_submitted_datetime(test_data: FixtureData):
    application = _create_wood_application("Wood ref 4", test_data)
    application.submit_datetime = make_aware(datetime.datetime(2020, 1, 1, 23, 59, 59))
    application.save()

    search_terms = SearchTerms(
        case_type="import",
        app_type=ImportApplicationType.Types.WOOD_QUOTA,  # type: ignore[arg-type]
        submitted_date_start=datetime.date(2020, 1, 2),
    )
    results = search_applications(search_terms)

    assert results.total_rows == 3

    check_application_references(results.records, "Wood ref 3", "Wood ref 2", "Wood ref 1")

    # Now search by end date to only find "Wood ref 4"
    search_terms = SearchTerms(
        case_type="import",
        app_type=ImportApplicationType.Types.WOOD_QUOTA,  # type: ignore[arg-type]
        submitted_date_start=datetime.date(2020, 1, 1),
        submitted_date_end=datetime.date(2020, 1, 2),
    )
    results = search_applications(search_terms)

    assert results.total_rows == 1

    check_application_references(results.records, "Wood ref 4")


def test_search_by_submitted_end_date(test_data: FixtureData):
    application = _create_wood_application("Wood ref 1", test_data)
    application.submit_datetime = make_aware(datetime.datetime(2020, 1, 2, 23, 59, 59))
    application.save()

    search_terms = SearchTerms(
        case_type="import",
        app_type=ImportApplicationType.Types.WOOD_QUOTA,  # type: ignore[arg-type]
        submitted_date_end=datetime.date(2020, 1, 2),
    )
    results = search_applications(search_terms)

    assert results.total_rows == 1

    check_application_references(results.records, "Wood ref 1")


def _test_search_by_licence_date():
    # Set the licence dates on a submitted application (26/AUG/2021 - 26/FEB/2022)
    application = WoodQuotaApplication.objects.get(applicant_reference="Wood ref 4")
    application.licence_start_date = datetime.date(2021, 8, 26)
    application.licence_end_date = datetime.date(2022, 2, 26)
    application.save()

    # Should find the record when the search terms are the same day as the licence dates
    search_terms = SearchTerms(
        case_type="import",
        app_type=ImportApplicationType.Types.WOOD_QUOTA,  # type: ignore[arg-type]
        licence_date_start=datetime.date(2021, 8, 26),
        licence_date_end=datetime.date(2022, 2, 26),
    )

    results = search_applications(search_terms)

    assert results.total_rows == 1

    check_application_references(results.records, "Wood ref 4")

    # A later start date should remove the above record
    search_terms = SearchTerms(
        case_type="import",
        app_type=ImportApplicationType.Types.WOOD_QUOTA,  # type: ignore[arg-type]
        licence_date_start=datetime.date(2021, 8, 27),
        licence_date_end=datetime.date(2022, 2, 26),
    )
    results = search_applications(search_terms)

    assert results.total_rows == 0

    # an earlier end date should remove the above record
    search_terms = SearchTerms(
        case_type="import",
        app_type=ImportApplicationType.Types.WOOD_QUOTA,  # type: ignore[arg-type]
        licence_date_start=datetime.date(2021, 8, 26),
        licence_date_end=datetime.date(2022, 2, 25),
    )
    results = search_applications(search_terms)

    assert results.total_rows == 0


def test_get_search_results_spreadsheet(test_data: FixtureData):
    _create_wood_application("Wood ref 1", test_data)
    _create_wood_application("Wood ref 2", test_data)
    _create_wood_application("Wood ref 3", test_data)
    _create_textiles_application("Textiles ref 1", test_data)
    _create_opt_application("Opt ref 1", test_data)
    _create_fa_dfl_application("fa-dfl ref 1", test_data)

    search_terms = SearchTerms(case_type="import")
    results = search_applications(search_terms)

    assert results.total_rows == 6
    check_application_references(
        results.records,
        "fa-dfl ref 1",
        "Opt ref 1",
        "Textiles ref 1",
        "Wood ref 3",
        "Wood ref 2",
        "Wood ref 1",
    )

    xlsx_data = get_search_results_spreadsheet(results)

    workbook = load_workbook(filename=io.BytesIO(xlsx_data))

    assert workbook.sheetnames == ["Sheet 1"]

    sheet = workbook["Sheet 1"]

    cols = sheet.max_column
    rows = sheet.max_row
    assert cols == 19
    assert rows == 7

    spreadsheet_rows = sheet.values
    header = next(spreadsheet_rows)

    assert list(header) == [
        "Case Reference",
        "Applicant's Reference",
        "Licence Reference",
        "Licence Type",
        "Licence Start Date",
        "Licence End Date",
        "Application Type",
        "Application Sub-Type",
        "Case Status",
        "Chief Usage Status",
        "Submitted Date",
        "Importer",
        "Agent",
        "Application Contact",
        "Country of Origin",
        "Country of Consignment",
        "Shipping Year",
        "Goods Category",
        "Commodity Code(s)",
    ]

    applicant_refs = []

    # Iterate over the remaining data rows
    for row in spreadsheet_rows:
        row_data = list(row)
        applicant_refs.append(row_data[1])

    assert applicant_refs == [
        "fa-dfl ref 1",
        "Opt ref 1",
        "Textiles ref 1",
        "Wood ref 3",
        "Wood ref 2",
        "Wood ref 1",
    ]


def test_case_statuses(test_data: FixtureData):
    _create_test_app_statuses(test_data)

    wt = ImportApplicationType.Types.WOOD_QUOTA
    st = ImportApplication.Statuses

    _test_search_by_status(wt, st.COMPLETED, expected=["completed"])

    # TODO: ICMSLST-1103 in progress doesn't return anything in v1
    # _test_search_by_status(wt, scs.IN_PROGRESS, [])

    # TODO: ICMSLST-1105 filter Oustanding Open Requests
    # _test_search_by_status(wt, scs.OPEN_REQUESTS, "open_request")

    _test_search_by_status(wt, st.PROCESSING, expected=["update", "fir", "processing"])
    _test_search_by_status(wt, st.FIR_REQUESTED, expected=["fir"])

    # TODO: ICMSLST-1104: filter SIGL
    # _test_search_by_status(wt, scs.PROCESSING_SIGL, "sigl")

    _test_search_by_status(wt, st.UPDATE_REQUESTED, expected=["update"])
    _test_search_by_status(wt, st.REVOKED, expected=["revoked"])
    _test_search_by_status(wt, st.STOPPED, expected=["stopped"])
    _test_search_by_status(wt, st.SUBMITTED, expected=["submitted"])
    _test_search_by_status(wt, st.VARIATION_REQUESTED, expected=["variation"])
    _test_search_by_status(wt, st.WITHDRAWN, expected=["withdrawn"])

    with pytest.raises(NotImplementedError):
        _test_search_by_status(wt, "unknown status", ["should raise"])


def test_search_by_export_applications(export_fixture_data: ExportFixtureData):
    cfs_app = _create_cfs_application(export_fixture_data)
    com_app = _create_com_application(export_fixture_data)
    gmp_app = _create_gmp_application(export_fixture_data)

    search_terms = SearchTerms(case_type="export")
    results = search_applications(search_terms)

    assert results.total_rows == 3
    check_export_application_case_reference(
        results.records, gmp_app.reference, com_app.reference, cfs_app.reference
    )


def test_search_by_app_type(export_fixture_data: ExportFixtureData):
    cfs_app = _create_cfs_application(export_fixture_data)
    com_app = _create_com_application(export_fixture_data)
    gmp_app = _create_gmp_application(export_fixture_data)

    ref_type_pairs = (
        (cfs_app.reference, ExportApplicationType.Types.FREE_SALE),
        (com_app.reference, ExportApplicationType.Types.MANUFACTURE),
        (gmp_app.reference, ExportApplicationType.Types.GMP),
    )

    for (ref, app_type) in ref_type_pairs:
        search_terms = SearchTerms(case_type="export", app_type=app_type)
        results = search_applications(search_terms)

        assert results.total_rows == 1, f"Failed: {ref} - {app_type}"
        check_export_application_case_reference(results.records, ref)


def test_export_search_by_exporter_or_agent(export_fixture_data: ExportFixtureData):
    cfs_app = _create_cfs_application(export_fixture_data)
    search_terms = SearchTerms(
        case_type="export", exporter_agent_name=export_fixture_data.exporter.name
    )
    results = search_applications(search_terms)

    assert results.total_rows == 1
    check_export_application_case_reference(results.records, cfs_app.reference)

    cfs_app = _create_cfs_application(export_fixture_data)
    cfs_app.agent = export_fixture_data.agent_exporter
    cfs_app.save()

    search_terms = SearchTerms(
        case_type="export", exporter_agent_name=export_fixture_data.agent_exporter.name
    )
    results = search_applications(search_terms)

    assert results.total_rows == 1
    check_export_application_case_reference(results.records, cfs_app.reference)

    search_terms = SearchTerms(case_type="export", exporter_agent_name="Not valid")
    results = search_applications(search_terms)

    assert results.total_rows == 0


def test_export_search_by_closed_dates(export_fixture_data: ExportFixtureData):
    # TODO: Add test when doing ICMSLST-1107
    ...


def test_export_search_by_certificate_country(export_fixture_data: ExportFixtureData):
    # all applications have the following certificate countries.
    # Aruba", "Maldives", "Zambia"

    aruba = Country.objects.filter(name="Aruba")
    yemen = Country.objects.filter(name="Yemen")
    aruba_and_yemen = Country.objects.filter(name__in=["Aruba", "Yemen"])

    cfs_app = _create_cfs_application(export_fixture_data)
    cfs_app2: CertificateOfFreeSaleApplication = _create_cfs_application(export_fixture_data)
    cfs_app2.countries.add(yemen.first())

    search_terms = SearchTerms(case_type="export", certificate_country=aruba)
    results = search_applications(search_terms)

    assert results.total_rows == 2
    check_export_application_case_reference(results.records, cfs_app2.reference, cfs_app.reference)

    search_terms = SearchTerms(case_type="export", certificate_country=yemen)
    results = search_applications(search_terms)

    assert results.total_rows == 1
    check_export_application_case_reference(results.records, cfs_app2.reference)

    search_terms = SearchTerms(case_type="export", certificate_country=aruba_and_yemen)
    results = search_applications(search_terms)

    assert results.total_rows == 2
    check_export_application_case_reference(results.records, cfs_app2.reference, cfs_app.reference)


def test_export_search_by_manufacture_country(export_fixture_data: ExportFixtureData):
    # Default country of manufacture is Peru
    app = _create_cfs_application(export_fixture_data)

    peru = Country.objects.filter(name="Peru")
    yemen = Country.objects.filter(name="Yemen")
    peru_and_yemen = Country.objects.filter(name__in=["Peru", "Yemen"])

    search_terms = SearchTerms(case_type="export", manufacture_country=peru)
    results = search_applications(search_terms)

    assert results.total_rows == 1
    check_export_application_case_reference(results.records, app.reference)

    search_terms = SearchTerms(case_type="export", manufacture_country=yemen)
    results = search_applications(search_terms)

    assert results.total_rows == 0

    search_terms = SearchTerms(case_type="export", manufacture_country=peru_and_yemen)
    results = search_applications(search_terms)

    assert results.total_rows == 1
    check_export_application_case_reference(results.records, app.reference)


def test_export_search_by_pending_firs(export_fixture_data: ExportFixtureData):
    com_app = _create_com_application(export_fixture_data)
    search_terms = SearchTerms(case_type="export", pending_firs=YesNoChoices.yes)
    results = search_applications(search_terms)

    assert results.total_rows == 0

    com_app.status = ApplicationBase.Statuses.FIR_REQUESTED
    com_app.save()

    results = search_applications(search_terms)

    assert results.total_rows == 1
    check_export_application_case_reference(results.records, com_app.reference)


def test_export_search_by_pending_update_reqs(export_fixture_data: ExportFixtureData):
    com_app = _create_com_application(export_fixture_data)
    search_terms = SearchTerms(case_type="export", pending_update_reqs=YesNoChoices.yes)
    results = search_applications(search_terms)

    assert results.total_rows == 0

    com_app.status = ApplicationBase.Statuses.UPDATE_REQUESTED
    com_app.save()

    results = search_applications(search_terms)

    assert results.total_rows == 1
    check_export_application_case_reference(results.records, com_app.reference)


def check_application_references(applications: list[ResultRow], *references, sort_results=False):
    """Check the returned applications match the supplied references

    Sort results if we don't care about the order
    """

    expected = sorted(references) if sort_results else list(references)

    actual = (app.case_status.applicant_reference for app in applications)
    actual = sorted(actual) if sort_results else list(actual)  # type: ignore[type-var, assignment]

    assert expected == actual


def check_export_application_case_reference(
    applications: list[ResultRow], *app_references, sort_results=False
):
    expected = sorted(app_references) if sort_results else list(app_references)

    actual = (app.case_reference for app in applications)
    actual = sorted(actual) if sort_results else list(actual)  # type: ignore[type-var, assignment]

    assert expected == actual


def check_commodity_details(
    actual_details: CommodityDetails,
    *,
    expected_origin_country: str = None,
    expected_consignment_country: str = None,
    expected_goods_category: str = None,
    expected_shipping_year: int = None,
    expected_commodity_codes: list[str] = None,
):

    assert expected_origin_country == actual_details.origin_country
    assert expected_consignment_country == actual_details.consignment_country
    assert expected_goods_category == actual_details.goods_category
    assert expected_shipping_year == actual_details.shipping_year
    assert expected_commodity_codes == actual_details.commodity_codes


def _create_derogation_application(
    reference,
    test_data: FixtureData,
    submit=True,
    origin_country="Tanzania",
    consignment_country="Algeria",
    commodity_code="code112233",
):

    application_type = ImportApplicationType.objects.get(
        type=ImportApplicationType.Types.DEROGATION
    )
    process_type = ImportApplicationType.ProcessTypes.DEROGATIONS.value  # type: ignore[attr-defined]
    commodity = _create_test_commodity(commodity_code)

    derogation_kwargs = {
        "origin_country": Country.objects.get(name=origin_country),
        "consignment_country": Country.objects.get(name=consignment_country),
        "commodity": commodity,
    }

    return _create_application(
        application_type, process_type, reference, test_data, submit, extra_kwargs=derogation_kwargs
    )


def _create_fa_dfl_application(
    reference,
    test_data: FixtureData,
    submit=True,
    origin_country="the Czech Republic",
    consignment_country="the Slovak Republic",
    commodity_code=FirearmCommodity.EX_CHAPTER_97,
):
    application_type = ImportApplicationType.objects.get(
        type=ImportApplicationType.Types.FIREARMS, sub_type=ImportApplicationType.SubTypes.DFL
    )
    process_type = ImportApplicationType.ProcessTypes.FA_DFL.value  # type: ignore[attr-defined]

    fa_dfl_kwargs = {
        "origin_country": Country.objects.get(name=origin_country),
        "consignment_country": Country.objects.get(name=consignment_country),
        "commodity_code": commodity_code,
    }

    return _create_application(
        application_type, process_type, reference, test_data, submit, extra_kwargs=fa_dfl_kwargs
    )


def _create_fa_oil_application(
    reference,
    test_data: FixtureData,
    submit=True,
    origin_country="Any Country",
    consignment_country="Any Country",
    commodity_code=FirearmCommodity.EX_CHAPTER_93,
):
    application_type = ImportApplicationType.objects.get(
        type=ImportApplicationType.Types.FIREARMS, sub_type=ImportApplicationType.SubTypes.OIL
    )
    process_type = ImportApplicationType.ProcessTypes.FA_OIL.value  # type: ignore[attr-defined]
    fa_oil_kwargs = {
        "origin_country": Country.objects.get(name=origin_country),
        "consignment_country": Country.objects.get(name=consignment_country),
        "commodity_code": commodity_code,
    }

    return _create_application(
        application_type, process_type, reference, test_data, submit, extra_kwargs=fa_oil_kwargs
    )


def _create_fa_sil_application(
    reference,
    test_data: FixtureData,
    submit=True,
    origin_country="Argentina",
    consignment_country="Azerbaijan",
    commodity_code=FirearmCommodity.EX_CHAPTER_97,
):
    application_type = ImportApplicationType.objects.get(
        type=ImportApplicationType.Types.FIREARMS, sub_type=ImportApplicationType.SubTypes.SIL
    )
    process_type = ImportApplicationType.ProcessTypes.FA_SIL.value  # type: ignore[attr-defined]
    fa_sil_kwargs = {
        "origin_country": Country.objects.get(name=origin_country),
        "consignment_country": Country.objects.get(name=consignment_country),
        "commodity_code": commodity_code,
    }

    return _create_application(
        application_type, process_type, reference, test_data, submit, extra_kwargs=fa_sil_kwargs
    )


def _create_ironsteel_application(
    reference,
    test_data: FixtureData,
    submit=True,
    origin_country="Kazakhstan",
    consignment_country="Bahamas",
    shipping_year=2021,
    category_commodity_group="SA1",
    commodity_code="7208370010",
):
    application_type = ImportApplicationType.objects.get(
        type=ImportApplicationType.Types.IRON_STEEL
    )
    process_type = ImportApplicationType.ProcessTypes.IRON_STEEL.value  # type: ignore[attr-defined]
    commodity = _create_test_commodity(commodity_code)
    commodity_group = _create_test_commodity_group(category_commodity_group, commodity)

    ironsteel_kwargs = {
        "origin_country": Country.objects.get(name=origin_country),
        "consignment_country": Country.objects.get(name=consignment_country),
        "shipping_year": shipping_year,
        "commodity": commodity,
        "category_commodity_group": commodity_group,
    }

    return _create_application(
        application_type, process_type, reference, test_data, submit, extra_kwargs=ironsteel_kwargs
    )


def _create_opt_application(
    reference,
    test_data: FixtureData,
    origin_country="Uruguay",
    consignment_country="USA",
    cp_category=CP_CATEGORIES[0],
    cp_commodity_codes=("6205200010", "6205908010"),
    teg_commodity_codes=("5006009000", "5007206190", "5112301000"),
):
    application_type = ImportApplicationType.objects.get(type=ImportApplicationType.Types.OPT)
    process_type = ImportApplicationType.ProcessTypes.OPT.value  # type: ignore[attr-defined]
    cp_commodities = []
    teg_commodities = []

    for cc in cp_commodity_codes:
        cp_commodities.append(_create_test_commodity(cc))

    for cc in teg_commodity_codes:
        teg_commodities.append(_create_test_commodity(cc))

    opt_kwargs = {
        "cp_origin_country": Country.objects.get(name=origin_country),
        "cp_processing_country": Country.objects.get(name=consignment_country),
        "cp_category": cp_category,
    }

    application: OutwardProcessingTradeApplication = _create_application(
        application_type, process_type, reference, test_data, submit=False, extra_kwargs=opt_kwargs
    )

    for com in cp_commodities:
        application.cp_commodities.add(com)

    for com in teg_commodities:
        application.teg_commodities.add(com)

    _submit_application(application, test_data)

    return application


def _create_sanctionadhoc_application(
    reference,
    test_data: FixtureData,
    origin_country="Iran",
    consignment_country="Algeria",
    commodity_codes=("2801000010", "2850002070"),
):
    application_type = ImportApplicationType.objects.get(
        type=ImportApplicationType.Types.SANCTION_ADHOC
    )
    process_type = ImportApplicationType.ProcessTypes.SANCTIONS.value  # type: ignore[attr-defined]

    sanctionadhoc_kwargs = {
        "origin_country": Country.objects.get(name=origin_country),
        "consignment_country": Country.objects.get(name=consignment_country),
    }

    application = _create_application(
        application_type,
        process_type,
        reference,
        test_data,
        submit=False,
        extra_kwargs=sanctionadhoc_kwargs,
    )

    for com in commodity_codes:
        SanctionsAndAdhocApplicationGoods.objects.create(
            import_application=application,
            commodity=_create_test_commodity(com),
            goods_description=f"Some goods: {com}",
            quantity_amount=123,
            value=123,
        )

    _submit_application(application, test_data)

    return application


def _create_sps_application(
    reference,
    test_data: FixtureData,
    submit=True,
    origin_country="Azerbaijan",
    consignment_country="Jordan",
    commodity_code="7208539000",
):
    application_type = ImportApplicationType.objects.get(type=ImportApplicationType.Types.SPS)
    process_type = ImportApplicationType.ProcessTypes.SPS.value  # type: ignore[attr-defined]

    sps_kwargs = {
        "origin_country": Country.objects.get(name=origin_country),
        "consignment_country": Country.objects.get(name=consignment_country),
        "commodity": _create_test_commodity(commodity_code),
    }

    return _create_application(
        application_type, process_type, reference, test_data, submit, extra_kwargs=sps_kwargs
    )


def _create_textiles_application(
    reference,
    test_data: FixtureData,
    submit=True,
    origin_country="Belarus",
    consignment_country="Argentina",
    shipping_year=2024,
    category_commodity_group="22",
    commodity_code="5509620000",
):
    application_type = ImportApplicationType.objects.get(type=ImportApplicationType.Types.TEXTILES)
    process_type = ImportApplicationType.ProcessTypes.TEXTILES.value  # type: ignore[attr-defined]

    commodity = _create_test_commodity(commodity_code)
    commodity_group = _create_test_commodity_group(category_commodity_group, commodity)

    textiles_kwargs = {
        "origin_country": Country.objects.get(name=origin_country),
        "consignment_country": Country.objects.get(name=consignment_country),
        "shipping_year": shipping_year,
        "commodity": commodity,
        "category_commodity_group": commodity_group,
    }

    return _create_application(
        application_type, process_type, reference, test_data, submit, extra_kwargs=textiles_kwargs
    )


def _create_wood_application(
    reference,
    test_data: FixtureData,
    submit=True,
    shipping_year=2021,
    commodity_code="code123456",
    override_status=None,
):
    application_type = ImportApplicationType.objects.get(
        type=ImportApplicationType.Types.WOOD_QUOTA
    )
    process_type = ImportApplicationType.ProcessTypes.WOOD.value  # type: ignore[attr-defined]
    commodity = _create_test_commodity(commodity_code)

    wood_kwargs = {
        "shipping_year": shipping_year,
        "commodity": commodity,
    }

    return _create_application(
        application_type,
        process_type,
        reference,
        test_data,
        submit,
        override_status,
        extra_kwargs=wood_kwargs,
    )


def _create_cfs_application(
    export_fixture_data: ExportFixtureData, submit=True, country_of_manufacture="Peru"
):
    application_type = ExportApplicationType.objects.get(
        type_code=ExportApplicationType.Types.FREE_SALE
    )
    process_type = ExportApplicationType.ProcessTypes.CFS.value

    com = Country.objects.get(name=country_of_manufacture)
    app: CertificateOfFreeSaleApplication = _create_export_application(
        application_type, process_type, export_fixture_data, False, extra_kwargs={}
    )

    app.schedules.create(country_of_manufacture=com, created_by=export_fixture_data.request.user)

    _submit_application(app, export_fixture_data)

    return app


def _create_com_application(export_fixture_data: ExportFixtureData, submit=True):
    application_type = ExportApplicationType.objects.get(
        type_code=ExportApplicationType.Types.MANUFACTURE
    )
    process_type = ExportApplicationType.ProcessTypes.COM.value

    return _create_export_application(
        application_type, process_type, export_fixture_data, submit, extra_kwargs={}
    )


def _create_gmp_application(export_fixture_data: ExportFixtureData, submit=True):
    application_type = ExportApplicationType.objects.get(type_code=ExportApplicationType.Types.GMP)
    process_type = ExportApplicationType.ProcessTypes.GMP.value

    return _create_export_application(
        application_type, process_type, export_fixture_data, submit, extra_kwargs={}
    )


def _create_application(
    application_type,
    process_type,
    reference,
    test_data,
    submit,
    override_status=None,
    extra_kwargs=None,
):
    kwargs = {
        "applicant_reference": reference,
        "importer": test_data.importer,
        "created_by": test_data.importer_user,
        "last_updated_by": test_data.importer_user,
        "submitted_by": test_data.importer_user,
        "application_type": application_type,
        "process_type": process_type,
        "contact": test_data.importer_user,
    }

    if extra_kwargs:
        kwargs.update(**extra_kwargs)

    models = {
        ImportApplicationType.ProcessTypes.DEROGATIONS: DerogationsApplication,
        ImportApplicationType.ProcessTypes.FA_DFL: DFLApplication,
        ImportApplicationType.ProcessTypes.FA_OIL: OpenIndividualLicenceApplication,
        ImportApplicationType.ProcessTypes.FA_SIL: SILApplication,
        ImportApplicationType.ProcessTypes.IRON_STEEL: IronSteelApplication,
        ImportApplicationType.ProcessTypes.OPT: OutwardProcessingTradeApplication,
        ImportApplicationType.ProcessTypes.SANCTIONS: SanctionsAndAdhocApplication,
        ImportApplicationType.ProcessTypes.SPS: PriorSurveillanceApplication,
        ImportApplicationType.ProcessTypes.TEXTILES: TextilesApplication,
        ImportApplicationType.ProcessTypes.WOOD: WoodQuotaApplication,
    }

    model_cls = models[process_type]

    application = model_cls.objects.create(**kwargs)  # type: ignore[attr-defined]
    Task.objects.create(
        process=application, task_type=Task.TaskType.PREPARE, owner=test_data.importer_user
    )

    if submit:
        _submit_application(application, test_data)

    if override_status:
        application.status = override_status
        application.save()

    return application


def _create_export_application(
    application_type: ExportApplicationType,
    process_type: str,
    fixture_data: ExportFixtureData,
    submit: bool,
    extra_kwargs: dict = None,
    certificate_countries=("Aruba", "Maldives", "Zambia"),
):
    kwargs = {
        # "applicant_reference": reference,
        "exporter": fixture_data.exporter,
        "created_by": fixture_data.exporter_user,
        "last_updated_by": fixture_data.exporter_user,
        "submitted_by": fixture_data.exporter_user,
        "application_type": application_type,
        "process_type": process_type,
        "contact": fixture_data.exporter_user,
    }

    if extra_kwargs:
        kwargs.update(**extra_kwargs)

    models = {
        ExportApplicationType.ProcessTypes.COM: CertificateOfManufactureApplication,
        ExportApplicationType.ProcessTypes.GMP: CertificateOfGoodManufacturingPracticeApplication,
        ExportApplicationType.ProcessTypes.CFS: CertificateOfFreeSaleApplication,
    }

    model_cls = models[process_type]

    application = model_cls.objects.create(**kwargs)  # type: ignore[attr-defined]
    Task.objects.create(
        process=application, task_type=Task.TaskType.PREPARE, owner=fixture_data.exporter_user
    )

    for c in certificate_countries:
        country = Country.objects.get(name=c)
        application.countries.add(country)

    application.save()

    if submit:
        _submit_application(application, fixture_data)

    return application


def _submit_application(application, test_data: Union[FixtureData, ExportFixtureData]):
    """Helper function to submit an application (Using the application code to do so)"""
    task = get_application_current_task(application, "import", Task.TaskType.PREPARE)

    application.submit_application(test_data.request, task)
    application.save()


def _create_test_commodity(commodity_code):
    com_type = CommodityType.objects.get(type_code="TEXTILES")
    commodity, created = Commodity.objects.get_or_create(
        defaults={"commodity_type": com_type, "validity_start_date": datetime.date.today()},
        commodity_code=commodity_code,
    )
    return commodity


def _create_test_commodity_group(category_commodity_group: str, commodity: Commodity):
    group, created = CommodityGroup.objects.get_or_create(group_code=category_commodity_group)

    if created:
        group.commodities.add(commodity)

    return group


def _create_test_app_statuses(test_data):
    st = ApplicationBase.Statuses
    _create_wood_application("completed", test_data, override_status=st.COMPLETED)

    # TODO: ICMSLST-1103 in progress doesn't return anything in v1
    # _create_wood_application("in_progress", test_data, submit=False, override_status=st.IN_PROGRESS)

    # TODO: ICMSLST-1105: filter Oustanding Open Requests
    # _create_wood_application("open_request", test_data)

    _create_wood_application("processing", test_data, override_status=st.PROCESSING)
    _create_wood_application("fir", test_data, override_status=st.FIR_REQUESTED)

    # TODO: ICMSLST-1104: filter SIGL
    # _create_wood_application("sigl", test_data)

    _create_wood_application("update", test_data, override_status=st.UPDATE_REQUESTED)
    _create_wood_application("revoked", test_data, override_status=st.REVOKED)
    _create_wood_application("stopped", test_data, override_status=st.STOPPED)
    _create_wood_application("submitted", test_data, override_status=st.SUBMITTED)
    _create_wood_application("variation", test_data, override_status=st.VARIATION_REQUESTED)
    _create_wood_application("withdrawn", test_data, override_status=st.WITHDRAWN)
