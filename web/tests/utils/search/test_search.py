import datetime
import io

import pytest
from django.urls import reverse
from django.utils.timezone import make_aware
from openpyxl import load_workbook

from web.domains.case._import.opt.models import CP_CATEGORIES
from web.domains.case.models import ApplicationBase, DocumentPackBase
from web.domains.case.services import document_pack
from web.domains.case.shared import ImpExpStatus
from web.models import (
    CaseEmail,
    CertificateOfFreeSaleApplication,
    CommodityGroup,
    Country,
    DFLChecklist,
    ExportApplicationType,
    FurtherInformationRequest,
    ImportApplication,
    ImportApplicationType,
    UpdateRequest,
    User,
    WoodQuotaApplication,
)
from web.models.shared import FirearmCommodity, YesNoChoices
from web.tests.helpers import CaseURLS
from web.utils.search import (
    SearchTerms,
    get_search_results_spreadsheet,
    search_applications,
    types,
)

from .conftest import (
    Build,
    ExportFixtureData,
    FixtureData,
    create_test_commodity,
    create_test_commodity_group,
)


def test_filter_by_application_type(importer_one_fixture_data):
    Build.wood_application("Wood ref 1", importer_one_fixture_data)
    Build.wood_application("Wood ref 2", importer_one_fixture_data)
    Build.derogation_application("Derogation ref 1", importer_one_fixture_data)
    Build.derogation_application("Derogation ref 2", importer_one_fixture_data)

    terms = SearchTerms(case_type="import", app_type=ImportApplicationType.Types.WOOD_QUOTA)
    results = search_applications(terms, importer_one_fixture_data.ilb_admin_user)

    assert results.total_rows == 2

    check_application_references(results.records, "Wood ref 2", "Wood ref 1")


def test_filter_wood(importer_one_fixture_data):
    """Do several tests related to searching for wood queries.

    All tests run in a single test for speed.
    The created applications are therefore reused in several tests.
    """

    _test_fetch_all(importer_one_fixture_data)

    _test_search_by_case_reference(importer_one_fixture_data)

    _test_search_by_applicant_reference(importer_one_fixture_data)

    _test_search_by_status(
        ImportApplicationType.Types.WOOD_QUOTA,
        ImpExpStatus.SUBMITTED,
        importer_one_fixture_data.ilb_admin_user,
        expected=["Wood ref 3", "Wood ref 2", "Wood ref 1"],
    )

    _test_search_by_response_decision(importer_one_fixture_data)

    _test_search_by_importer_or_agent_name(importer_one_fixture_data)

    _test_search_by_submitted_datetime(importer_one_fixture_data)

    _test_search_by_licence_date(importer_one_fixture_data)


def test_order_and_limit_works(importer_one_fixture_data):
    """Create many applications and ensure only the latest n submitted are returned"""

    applications = (
        "wood app 1",
        "wood app 2",
        "derogation app 1",
        "derogation app 2",
        "wood app 3",
        "wood app 4",
        "wood app 5",
        "derogation app 3",
        "derogation app 4",
        "derogation app 5",
    )

    for app_ref in applications:
        if app_ref.startswith("wood"):
            Build.wood_application(app_ref, importer_one_fixture_data)

        elif app_ref.startswith("derogation"):
            Build.derogation_application(app_ref, importer_one_fixture_data)

        else:
            raise Exception(f"failed to create: {app_ref}")

    terms = SearchTerms(case_type="import")
    search_data = search_applications(terms, importer_one_fixture_data.ilb_admin_user, limit=5)

    assert search_data.total_rows == 10

    assert len(search_data.records) == 5

    check_application_references(
        search_data.records,
        "derogation app 5",
        "derogation app 4",
        "derogation app 3",
        "wood app 5",
        "wood app 4",
    )


def test_derogation_commodity_details_correct(importer_one_fixture_data):
    app = Build.derogation_application("derogation app 1", importer_one_fixture_data)

    search_terms = SearchTerms(case_type="import", app_type=ImportApplicationType.Types.DEROGATION)
    results = search_applications(search_terms, importer_one_fixture_data.ilb_admin_user)

    assert results.total_rows == 1
    check_application_references(results.records, "derogation app 1")
    check_commodity_details(
        results.records[0].commodity_details,
        expected_origin_country="Tanzania",
        expected_consignment_country="Algeria",
        expected_shipping_year=app.submit_datetime.year,
        expected_commodity_codes=["code112233"],
    )


def test_fa_dfl_commodity_details_correct(importer_one_fixture_data):
    Build.fa_dfl_application("fa-dfl app 1", importer_one_fixture_data)

    search_terms = SearchTerms(
        case_type="import",
        app_type=ImportApplicationType.Types.FIREARMS,
        app_sub_type=ImportApplicationType.SubTypes.DFL,
    )
    results = search_applications(search_terms, importer_one_fixture_data.ilb_admin_user)

    assert results.total_rows == 1
    check_application_references(results.records, "fa-dfl app 1")
    check_commodity_details(
        results.records[0].commodity_details,
        expected_origin_country="the Czech Republic",
        expected_consignment_country="the Slovak Republic",
        expected_goods_category=FirearmCommodity.EX_CHAPTER_97.label,
    )


def test_fa_oil_commodity_details_correct(importer_one_fixture_data):
    Build.fa_oil_application("fa-oil app 1", importer_one_fixture_data)

    search_terms = SearchTerms(
        case_type="import",
        app_type=ImportApplicationType.Types.FIREARMS,
        app_sub_type=ImportApplicationType.SubTypes.OIL,
    )
    results = search_applications(search_terms, importer_one_fixture_data.ilb_admin_user)

    assert results.total_rows == 1
    check_application_references(results.records, "fa-oil app 1")
    check_commodity_details(
        results.records[0].commodity_details,
        expected_origin_country="Any Country",
        expected_consignment_country="Any Country",
        expected_goods_category=FirearmCommodity.EX_CHAPTER_93.label,
    )


def test_fa_sil_commodity_details_correct(importer_one_fixture_data):
    Build.fa_sil_application("fa-sil app 1", importer_one_fixture_data)

    search_terms = SearchTerms(
        case_type="import",
        app_type=ImportApplicationType.Types.FIREARMS,
        app_sub_type=ImportApplicationType.SubTypes.SIL,
    )
    results = search_applications(search_terms, importer_one_fixture_data.ilb_admin_user)

    assert results.total_rows == 1
    check_application_references(results.records, "fa-sil app 1")
    check_commodity_details(
        results.records[0].commodity_details,
        expected_origin_country="Argentina",
        expected_consignment_country="Azerbaijan",
        expected_goods_category=FirearmCommodity.EX_CHAPTER_97.label,
    )


def test_ironsteel_commodity_details_correct(importer_one_fixture_data):
    Build.ironsteel_application("ironsteel app 1", importer_one_fixture_data)

    search_terms = SearchTerms(case_type="import", app_type=ImportApplicationType.Types.IRON_STEEL)
    results = search_applications(search_terms, importer_one_fixture_data.ilb_admin_user)

    assert results.total_rows == 1
    check_application_references(results.records, "ironsteel app 1")
    check_commodity_details(
        results.records[0].commodity_details,
        expected_origin_country="Kazakhstan",
        expected_consignment_country="Bahamas",
        expected_shipping_year=2021,
        expected_goods_category="SA1",
        expected_commodity_codes=["7208370010"],
    )


def test_opt_commodity_details_correct(importer_one_fixture_data):
    app = Build.opt_application("opt app 1", importer_one_fixture_data)

    search_terms = SearchTerms(case_type="import", app_type=ImportApplicationType.Types.OPT)
    results = search_applications(search_terms, importer_one_fixture_data.ilb_admin_user)

    assert results.total_rows == 1
    check_application_references(results.records, "opt app 1")
    check_commodity_details(
        results.records[0].commodity_details,
        expected_origin_country="Uruguay",
        expected_consignment_country="USA",
        expected_shipping_year=app.submit_datetime.year,
        expected_goods_category=CP_CATEGORIES[0],
        expected_commodity_codes=[
            "5006009000",
            "5007206190",
            "5112301000",
            "6205200010",
            "6205908010",
        ],
    )


def test_opt_app_with_no_cp_commodities(importer_one_fixture_data):
    app = Build.opt_application("opt app 1", importer_one_fixture_data, cp_commodity_codes=[])

    search_terms = SearchTerms(case_type="import", app_type=ImportApplicationType.Types.OPT)
    results = search_applications(search_terms, importer_one_fixture_data.ilb_admin_user)

    assert results.total_rows == 1
    check_application_references(results.records, "opt app 1")
    check_commodity_details(
        results.records[0].commodity_details,
        expected_origin_country="Uruguay",
        expected_consignment_country="USA",
        expected_shipping_year=app.submit_datetime.year,
        expected_goods_category=CP_CATEGORIES[0],
        expected_commodity_codes=["5006009000", "5007206190", "5112301000"],
    )


def test_opt_app_with_no_teg_commodities(importer_one_fixture_data):
    app = Build.opt_application("opt app 1", importer_one_fixture_data, teg_commodity_codes=[])

    search_terms = SearchTerms(case_type="import", app_type=ImportApplicationType.Types.OPT)
    results = search_applications(search_terms, importer_one_fixture_data.ilb_admin_user)

    assert results.total_rows == 1
    check_application_references(results.records, "opt app 1")
    check_commodity_details(
        results.records[0].commodity_details,
        expected_origin_country="Uruguay",
        expected_consignment_country="USA",
        expected_shipping_year=app.submit_datetime.year,
        expected_goods_category=CP_CATEGORIES[0],
        expected_commodity_codes=["6205200010", "6205908010"],
    )


def test_sanctionadhoc_commodity_details_correct(importer_one_fixture_data):
    app = Build.sanctionadhoc_application("sanctionsadhoc app 1", importer_one_fixture_data)

    search_terms = SearchTerms(
        case_type="import", app_type=ImportApplicationType.Types.SANCTION_ADHOC
    )
    results = search_applications(search_terms, importer_one_fixture_data.ilb_admin_user)

    assert results.total_rows == 1
    check_application_references(results.records, "sanctionsadhoc app 1")
    check_commodity_details(
        results.records[0].commodity_details,
        expected_origin_country="Iran",
        expected_consignment_country="Algeria",
        expected_shipping_year=app.submit_datetime.year,
        expected_commodity_codes=["2801000010", "2850002070"],
    )


def test_sps_commodity_details_correct(importer_one_fixture_data):
    app = Build.sps_application("sps app 1", importer_one_fixture_data)

    search_terms = SearchTerms(case_type="import", app_type=ImportApplicationType.Types.SPS)
    results = search_applications(search_terms, importer_one_fixture_data.ilb_admin_user)

    assert results.total_rows == 1
    check_application_references(results.records, "sps app 1")
    check_commodity_details(
        results.records[0].commodity_details,
        expected_origin_country="Azerbaijan",
        expected_consignment_country="Jordan",
        expected_shipping_year=app.submit_datetime.year,
        expected_commodity_codes=["7208539000"],
    )


def test_sps_commodity_details_correct_multiple(importer_one_fixture_data):
    app_1 = Build.sps_application(
        "sps app 1",
        importer_one_fixture_data,
        origin_country="Afghanistan",
        consignment_country="Armenia",
        commodity_code="111111",
    )
    app_2 = Build.sps_application(
        "sps app 2",
        importer_one_fixture_data,
        origin_country="Albania",
        consignment_country="Australia",
        commodity_code="222222",
    )
    app_3 = Build.sps_application(
        "sps app 3",
        importer_one_fixture_data,
        origin_country="Algeria",
        consignment_country="Azerbaijan",
        commodity_code="333333",
    )
    app_4 = Build.sps_application(
        "sps app 4",
        importer_one_fixture_data,
        origin_country="Angola",
        consignment_country="Bahamas",
        commodity_code="444444",
    )
    app_5 = Build.sps_application(
        "sps app 5",
        importer_one_fixture_data,
        origin_country="Argentina",
        consignment_country="Bahrain",
        commodity_code="555555",
    )

    search_terms = SearchTerms(case_type="import", app_type=ImportApplicationType.Types.SPS)
    results = search_applications(search_terms, importer_one_fixture_data.ilb_admin_user)

    assert results.total_rows == 5
    check_application_references(
        results.records, "sps app 5", "sps app 4", "sps app 3", "sps app 2", "sps app 1"
    )

    check_commodity_details(
        results.records[4].commodity_details,
        expected_origin_country="Afghanistan",
        expected_consignment_country="Armenia",
        expected_shipping_year=app_1.submit_datetime.year,
        expected_commodity_codes=["111111"],
    )

    check_commodity_details(
        results.records[3].commodity_details,
        expected_origin_country="Albania",
        expected_consignment_country="Australia",
        expected_shipping_year=app_2.submit_datetime.year,
        expected_commodity_codes=["222222"],
    )

    check_commodity_details(
        results.records[2].commodity_details,
        expected_origin_country="Algeria",
        expected_consignment_country="Azerbaijan",
        expected_shipping_year=app_3.submit_datetime.year,
        expected_commodity_codes=["333333"],
    )

    check_commodity_details(
        results.records[1].commodity_details,
        expected_origin_country="Angola",
        expected_consignment_country="Bahamas",
        expected_shipping_year=app_4.submit_datetime.year,
        expected_commodity_codes=["444444"],
    )

    check_commodity_details(
        results.records[0].commodity_details,
        expected_origin_country="Argentina",
        expected_consignment_country="Bahrain",
        expected_shipping_year=app_5.submit_datetime.year,
        expected_commodity_codes=["555555"],
    )


def test_textiles_commodity_details_correct(importer_one_fixture_data):
    Build.textiles_application("textiles app 1", importer_one_fixture_data)

    search_terms = SearchTerms(case_type="import", app_type=ImportApplicationType.Types.TEXTILES)
    results = search_applications(search_terms, importer_one_fixture_data.ilb_admin_user)

    assert results.total_rows == 1
    check_application_references(results.records, "textiles app 1")
    check_commodity_details(
        results.records[0].commodity_details,
        expected_origin_country="Belarus",
        expected_consignment_country="Argentina",
        expected_shipping_year=2024,
        expected_goods_category="22",
        expected_commodity_codes=["5509620000"],
    )


def test_wood_commodity_details_correct(importer_one_fixture_data):
    Build.wood_application(
        "Wood ref 1", importer_one_fixture_data, shipping_year=2030, commodity_code="code654321"
    )

    search_terms = SearchTerms(case_type="import", app_type=ImportApplicationType.Types.WOOD_QUOTA)
    results = search_applications(search_terms, importer_one_fixture_data.ilb_admin_user)

    assert results.total_rows == 1

    check_application_references(results.records, "Wood ref 1")

    wood_app = results.records[0]

    assert wood_app.commodity_details.origin_country == "None"
    assert wood_app.commodity_details.shipping_year == 2030
    assert wood_app.commodity_details.commodity_codes == ["code654321"]


def _test_fetch_all(importer_conf: FixtureData):
    Build.wood_application("Wood in progress", importer_conf, submit=False)
    Build.wood_application("Wood ref 1", importer_conf)
    Build.wood_application("Wood ref 2", importer_conf)

    search_terms = SearchTerms(case_type="import", app_type=ImportApplicationType.Types.WOOD_QUOTA)
    results = search_applications(search_terms, importer_conf.request.user)

    assert results.total_rows == 2

    check_application_references(results.records, "Wood ref 2", "Wood ref 1")


def _test_search_by_case_reference(importer_conf: FixtureData):
    """Test submitting an application and searching for it by the case reference"""

    application = Build.wood_application("Wood ref 3", importer_conf)
    case_reference = application.get_reference()

    assert case_reference != "Not Assigned"

    search_terms = SearchTerms(
        case_type="import", app_type=ImportApplicationType.Types.WOOD_QUOTA, case_ref=case_reference
    )
    results = search_applications(search_terms, importer_conf.request.user)

    assert results.total_rows == 1
    check_application_references(results.records, "Wood ref 3")


def _test_search_by_applicant_reference(importer_conf: FixtureData):
    """We have Wood ref 1, 2 and 3 when this test is run."""

    search_terms = SearchTerms(case_type="import", applicant_ref="wood ref %")
    results = search_applications(search_terms, importer_conf.request.user)

    assert results.total_rows == 3
    check_application_references(results.records, "Wood ref 3", "Wood ref 2", "Wood ref 1")

    search_terms.applicant_ref = "Wood % 3"
    results = search_applications(search_terms, importer_conf.request.user)

    assert results.total_rows == 1
    check_application_references(results.records, "Wood ref 3")


def _test_search_by_status(app_type: str, case_status: str, user: User, expected: list[str]):
    """Search by status using the records we have already created"""

    search_terms = SearchTerms(case_type="import", app_type=app_type, case_status=case_status)
    results = search_applications(search_terms, user)

    assert results.total_rows == len(expected)
    check_application_references(results.records, *expected)


def _test_export_search_by_status(app_type: str, case_status: str, user: User, expected: list[str]):
    """Search by status using the records we have already created"""

    search_terms = SearchTerms(case_type="export", app_type=app_type, case_status=case_status)
    results = search_applications(search_terms, user)

    assert results.total_rows == len(expected)
    check_export_application_case_reference(results.records, *expected)


def _test_search_by_response_decision(importer_conf: FixtureData):
    submitted_application = WoodQuotaApplication.objects.get(applicant_reference="Wood ref 3")
    submitted_application.decision = ApplicationBase.APPROVE
    submitted_application.save()

    search_terms = SearchTerms(
        case_type="import",
        app_type=ImportApplicationType.Types.WOOD_QUOTA,
        response_decision=ApplicationBase.REFUSE,
    )
    results = search_applications(search_terms, importer_conf.request.user)

    assert results.total_rows == 0

    search_terms = SearchTerms(
        case_type="import",
        app_type=ImportApplicationType.Types.WOOD_QUOTA,
        response_decision=ApplicationBase.APPROVE,
    )
    results = search_applications(search_terms, importer_conf.request.user)

    assert results.total_rows == 1

    check_application_references(results.records, "Wood ref 3")


def _test_search_by_importer_or_agent_name(importer_conf: FixtureData):
    search_terms = SearchTerms(
        case_type="import",
        app_type=ImportApplicationType.Types.WOOD_QUOTA,
        importer_agent_name="Not valid",
    )
    results = search_applications(search_terms, importer_conf.request.user)

    assert results.total_rows == 0

    search_terms = SearchTerms(
        case_type="import",
        app_type=ImportApplicationType.Types.WOOD_QUOTA,
        importer_agent_name=importer_conf.importer.name,
    )
    results = search_applications(search_terms, importer_conf.request.user)

    assert results.total_rows == 3

    check_application_references(results.records, "Wood ref 3", "Wood ref 2", "Wood ref 1")

    # Set an agent on the first application and check we can search for that.
    application = WoodQuotaApplication.objects.get(applicant_reference="Wood ref 1")
    application.agent = importer_conf.agent_importer
    application.save()

    search_terms = SearchTerms(
        case_type="import",
        app_type=ImportApplicationType.Types.WOOD_QUOTA,
        importer_agent_name=importer_conf.agent_importer.name,
    )
    results = search_applications(search_terms, importer_conf.request.user)

    assert results.total_rows == 1

    check_application_references(results.records, "Wood ref 1")


def _test_search_by_submitted_datetime(importer_conf: FixtureData):
    application = Build.wood_application("Wood ref 4", importer_conf)
    application.submit_datetime = make_aware(datetime.datetime(2020, 1, 1, 23, 59, 59))
    application.save()

    search_terms = SearchTerms(
        case_type="import",
        app_type=ImportApplicationType.Types.WOOD_QUOTA,
        submitted_date_start=datetime.date(2020, 1, 2),
    )
    results = search_applications(search_terms, importer_conf.request.user)

    assert results.total_rows == 3

    check_application_references(results.records, "Wood ref 3", "Wood ref 2", "Wood ref 1")

    # Now search by end date to only find "Wood ref 4"
    search_terms = SearchTerms(
        case_type="import",
        app_type=ImportApplicationType.Types.WOOD_QUOTA,
        submitted_date_start=datetime.date(2020, 1, 1),
        submitted_date_end=datetime.date(2020, 1, 2),
    )
    results = search_applications(search_terms, importer_conf.request.user)

    assert results.total_rows == 1

    check_application_references(results.records, "Wood ref 4")


def test_search_by_submitted_end_date(importer_one_fixture_data):
    application = Build.wood_application("Wood ref 1", importer_one_fixture_data)
    application.submit_datetime = make_aware(datetime.datetime(2020, 1, 2, 23, 59, 59))
    application.save()

    search_terms = SearchTerms(
        case_type="import",
        app_type=ImportApplicationType.Types.WOOD_QUOTA,
        submitted_date_end=datetime.date(2020, 1, 2),
    )
    results = search_applications(search_terms, importer_one_fixture_data.ilb_admin_user)

    assert results.total_rows == 1

    check_application_references(results.records, "Wood ref 1")


def _test_search_by_licence_date(importer_conf: FixtureData):
    # Set the licence dates on a submitted application (26/AUG/2021 - 26/FEB/2022)
    application = WoodQuotaApplication.objects.get(applicant_reference="Wood ref 4")
    application.licences.create(
        licence_start_date=datetime.date(2021, 8, 26), licence_end_date=datetime.date(2022, 2, 26)
    )

    # Should find the record when the search terms are the same day as the licence dates
    search_terms = SearchTerms(
        case_type="import",
        app_type=ImportApplicationType.Types.WOOD_QUOTA,
        licence_date_start=datetime.date(2021, 8, 26),
        licence_date_end=datetime.date(2022, 2, 26),
    )

    results = search_applications(search_terms, importer_conf.request.user)

    assert results.total_rows == 1

    check_application_references(results.records, "Wood ref 4")

    # A later start date should remove the above record
    search_terms = SearchTerms(
        case_type="import",
        app_type=ImportApplicationType.Types.WOOD_QUOTA,
        licence_date_start=datetime.date(2021, 8, 27),
        licence_date_end=datetime.date(2022, 2, 26),
    )
    results = search_applications(search_terms, importer_conf.request.user)

    assert results.total_rows == 0

    # an earlier end date should remove the above record
    search_terms = SearchTerms(
        case_type="import",
        app_type=ImportApplicationType.Types.WOOD_QUOTA,
        licence_date_start=datetime.date(2021, 8, 26),
        licence_date_end=datetime.date(2022, 2, 25),
    )
    results = search_applications(search_terms, importer_conf.request.user)

    assert results.total_rows == 0


def test_get_search_results_spreadsheet(importer_one_fixture_data):
    Build.wood_application("Wood ref 1", importer_one_fixture_data)
    Build.wood_application("Wood ref 2", importer_one_fixture_data)
    Build.wood_application("Wood ref 3", importer_one_fixture_data)
    Build.textiles_application("Textiles ref 1", importer_one_fixture_data)
    Build.opt_application("Opt ref 1", importer_one_fixture_data)
    Build.fa_dfl_application("fa-dfl ref 1", importer_one_fixture_data)

    search_terms = SearchTerms(case_type="import")
    results = search_applications(search_terms, importer_one_fixture_data.ilb_admin_user)

    assert results.total_rows == 6
    check_application_references(
        results.records,
        "fa-dfl ref 1",
        "Opt ref 1",
        "Textiles ref 1",
        "Wood ref 3",
        "Wood ref 2",
        "Wood ref 1",
    )

    xlsx_data = get_search_results_spreadsheet("import", results)

    workbook = load_workbook(filename=io.BytesIO(xlsx_data))

    assert workbook.sheetnames == ["Sheet 1"]

    sheet = workbook["Sheet 1"]

    cols = sheet.max_column
    rows = sheet.max_row
    assert cols == 19
    assert rows == 7

    spreadsheet_rows = sheet.values
    header = next(spreadsheet_rows)

    assert list(header) == [
        "Case Reference",
        "Applicant's Reference",
        "Licence Reference",
        "Licence Type",
        "Licence Start Date",
        "Licence End Date",
        "Application Type",
        "Application Sub-Type",
        "Case Status",
        "Chief Usage Status",
        "Submitted Date",
        "Importer",
        "Agent",
        "Application Contact",
        "Country of Origin",
        "Country of Consignment",
        "Shipping Year",
        "Goods Category",
        "Commodity Code(s)",
    ]

    applicant_refs = []

    # Iterate over the remaining data rows
    for row_data in spreadsheet_rows:
        applicant_refs.append(row_data[1])

    assert applicant_refs == [
        "fa-dfl ref 1",
        "Opt ref 1",
        "Textiles ref 1",
        "Wood ref 3",
        "Wood ref 2",
        "Wood ref 1",
    ]


def test_get_export_search_results_spreadsheet(exporter_one_fixture_data: ExportFixtureData):
    gmp = Build.gmp_application(exporter_one_fixture_data)
    cfs = Build.cfs_application(exporter_one_fixture_data)
    com = Build.com_application(exporter_one_fixture_data)

    search_terms = SearchTerms(case_type="export")
    results = search_applications(search_terms, exporter_one_fixture_data.ilb_admin_user)

    assert results.total_rows == 3

    check_export_application_case_reference(
        results.records, com.reference, cfs.reference, gmp.reference
    )

    xlsx_data = get_search_results_spreadsheet("export", results)

    workbook = load_workbook(filename=io.BytesIO(xlsx_data))

    assert workbook.sheetnames == ["Sheet 1"]
    sheet = workbook["Sheet 1"]

    cols = sheet.max_column
    rows = sheet.max_row
    assert cols == 10
    assert rows == 4

    spreadsheet_rows = sheet.values
    header = next(spreadsheet_rows)

    assert list(header) == [
        "Case Reference",
        "Certificates",
        "Application Type",
        "Status",
        "Submitted Date",
        "Certificate Countries",
        "Countries of Manufacture",
        "Exporter",
        "Agent",
        "Application Contact",
    ]

    case_refs = []

    # Iterate over the remaining data rows
    for row_data in spreadsheet_rows:
        case_refs.append(row_data[0])

    assert case_refs == [com.reference, cfs.reference, gmp.reference]


def test_case_statuses(importer_one_fixture_data):
    st = ImpExpStatus
    Build.wood_application("completed", importer_one_fixture_data, override_status=st.COMPLETED)

    # TODO: ICMSLST-1105: filter Oustanding Open Requests
    # Build.wood_application("open_request", importer_one_fixture_data)

    Build.wood_application("processing", importer_one_fixture_data, override_status=st.PROCESSING)
    app = Build.wood_application("fir", importer_one_fixture_data, override_status=st.PROCESSING)
    app.further_information_requests.create(
        status=FurtherInformationRequest.OPEN, process_type=FurtherInformationRequest.PROCESS_TYPE
    )

    # TODO: ICMSLST-1104: filter SIGL
    # Build.wood_application("sigl", importer_one_fixture_data)

    app = Build.wood_application("update", importer_one_fixture_data, override_status=st.PROCESSING)
    app.update_requests.create(status=UpdateRequest.Status.OPEN)

    Build.wood_application("revoked", importer_one_fixture_data, override_status=st.REVOKED)
    Build.wood_application("stopped", importer_one_fixture_data, override_status=st.STOPPED)
    Build.wood_application("submitted", importer_one_fixture_data, override_status=st.SUBMITTED)
    Build.wood_application(
        "variation", importer_one_fixture_data, override_status=st.VARIATION_REQUESTED
    )
    Build.wood_application("withdrawn", importer_one_fixture_data, override_status=st.WITHDRAWN)

    wt = ImportApplicationType.Types.WOOD_QUOTA
    user = importer_one_fixture_data.ilb_admin_user
    _test_search_by_status(wt, ImpExpStatus.COMPLETED, user, expected=["completed"])

    # TODO: ICMSLST-1105 filter Oustanding Open Requests
    # _test_search_by_status(wt, scs.OPEN_REQUESTS, "open_request")

    _test_search_by_status(
        wt, ImpExpStatus.PROCESSING, user, expected=["update", "fir", "processing"]
    )
    _test_search_by_status(wt, "FIR_REQUESTED", user, expected=["fir"])

    # TODO: ICMSLST-1104: filter SIGL
    # _test_search_by_status(wt, scs.PROCESSING_SIGL, "sigl")

    _test_search_by_status(wt, "UPDATE_REQUESTED", user, expected=["update"])
    _test_search_by_status(wt, ImpExpStatus.REVOKED, user, expected=["revoked"])
    _test_search_by_status(wt, ImpExpStatus.STOPPED, user, expected=["stopped"])
    _test_search_by_status(wt, ImpExpStatus.SUBMITTED, user, expected=["submitted"])
    _test_search_by_status(wt, ImpExpStatus.VARIATION_REQUESTED, user, expected=["variation"])
    _test_search_by_status(wt, ImpExpStatus.WITHDRAWN, user, expected=["withdrawn"])

    with pytest.raises(NotImplementedError):
        _test_search_by_status(wt, "unknown status", user, ["should raise"])


def test_export_case_statuses(exporter_one_fixture_data: ExportFixtureData):
    gmp = Build.gmp_application(exporter_one_fixture_data)
    gmp.status = ImpExpStatus.PROCESSING
    gmp.save()
    gmp.case_emails.create(status=CaseEmail.Status.OPEN)

    cfs = Build.cfs_application(exporter_one_fixture_data)
    cfs.status = ImpExpStatus.PROCESSING
    cfs.save()
    cfs.case_emails.create(status=CaseEmail.Status.OPEN)

    user = exporter_one_fixture_data.ilb_admin_user

    _test_export_search_by_status(
        ExportApplicationType.Types.GMP, "BEIS", user, expected=[gmp.reference]
    )
    _test_export_search_by_status(
        ExportApplicationType.Types.FREE_SALE, "HSE", user, expected=[cfs.reference]
    )


def test_search_by_export_applications(exporter_one_fixture_data: ExportFixtureData):
    cfs_app = Build.cfs_application(exporter_one_fixture_data)
    com_app = Build.com_application(exporter_one_fixture_data)
    gmp_app = Build.gmp_application(exporter_one_fixture_data)

    search_terms = SearchTerms(case_type="export")
    results = search_applications(search_terms, exporter_one_fixture_data.ilb_admin_user)

    assert results.total_rows == 3
    check_export_application_case_reference(
        results.records, gmp_app.reference, com_app.reference, cfs_app.reference
    )


def test_search_by_app_type(exporter_one_fixture_data: ExportFixtureData):
    cfs_app = Build.cfs_application(exporter_one_fixture_data)
    com_app = Build.com_application(exporter_one_fixture_data)
    gmp_app = Build.gmp_application(exporter_one_fixture_data)

    ref_type_pairs = (
        (cfs_app.reference, ExportApplicationType.Types.FREE_SALE),
        (com_app.reference, ExportApplicationType.Types.MANUFACTURE),
        (gmp_app.reference, ExportApplicationType.Types.GMP),
    )

    for ref, app_type in ref_type_pairs:
        search_terms = SearchTerms(case_type="export", app_type=app_type)
        results = search_applications(search_terms, exporter_one_fixture_data.ilb_admin_user)

        assert results.total_rows == 1, f"Failed: {ref} - {app_type}"
        check_export_application_case_reference(results.records, ref)


def test_export_search_by_exporter_or_agent(exporter_one_fixture_data: ExportFixtureData):
    cfs_app = Build.cfs_application(exporter_one_fixture_data)
    search_terms = SearchTerms(
        case_type="export", exporter_agent_name=exporter_one_fixture_data.exporter.name
    )
    results = search_applications(search_terms, exporter_one_fixture_data.ilb_admin_user)

    assert results.total_rows == 1
    check_export_application_case_reference(results.records, cfs_app.reference)

    cfs_app = Build.cfs_application(exporter_one_fixture_data)
    cfs_app.agent = exporter_one_fixture_data.agent_exporter
    cfs_app.save()

    search_terms = SearchTerms(
        case_type="export", exporter_agent_name=exporter_one_fixture_data.agent_exporter.name
    )
    results = search_applications(search_terms, exporter_one_fixture_data.ilb_admin_user)

    assert results.total_rows == 1
    check_export_application_case_reference(results.records, cfs_app.reference)

    search_terms = SearchTerms(case_type="export", exporter_agent_name="Not valid")
    results = search_applications(search_terms, exporter_one_fixture_data.ilb_admin_user)

    assert results.total_rows == 0


def test_export_search_by_closed_dates(exporter_one_fixture_data: ExportFixtureData):
    # TODO: Add test when doing ICMSLST-1107
    ...


def test_export_search_by_certificate_country(exporter_one_fixture_data: ExportFixtureData):
    # all applications have the following certificate countries.
    # Aruba", "Maldives", "Zambia"

    aruba = Country.objects.filter(name="Aruba")
    yemen = Country.objects.filter(name="Yemen")
    aruba_and_yemen = Country.objects.filter(name__in=["Aruba", "Yemen"])

    cfs_app = Build.cfs_application(exporter_one_fixture_data)
    cfs_app2: CertificateOfFreeSaleApplication = Build.cfs_application(exporter_one_fixture_data)
    cfs_app2.countries.add(yemen.first())

    search_terms = SearchTerms(case_type="export", certificate_country=aruba)
    results = search_applications(search_terms, exporter_one_fixture_data.ilb_admin_user)

    assert results.total_rows == 2
    check_export_application_case_reference(results.records, cfs_app2.reference, cfs_app.reference)

    search_terms = SearchTerms(case_type="export", certificate_country=yemen)
    results = search_applications(search_terms, exporter_one_fixture_data.ilb_admin_user)

    assert results.total_rows == 1
    check_export_application_case_reference(results.records, cfs_app2.reference)

    search_terms = SearchTerms(case_type="export", certificate_country=aruba_and_yemen)
    results = search_applications(search_terms, exporter_one_fixture_data.ilb_admin_user)

    assert results.total_rows == 2
    check_export_application_case_reference(results.records, cfs_app2.reference, cfs_app.reference)


def test_export_search_by_manufacture_country(exporter_one_fixture_data: ExportFixtureData):
    # Default country of manufacture is Peru
    app = Build.cfs_application(exporter_one_fixture_data)

    peru = Country.objects.filter(name="Peru")
    yemen = Country.objects.filter(name="Yemen")
    peru_and_yemen = Country.objects.filter(name__in=["Peru", "Yemen"])

    search_terms = SearchTerms(case_type="export", manufacture_country=peru)
    results = search_applications(search_terms, exporter_one_fixture_data.ilb_admin_user)

    assert results.total_rows == 1
    check_export_application_case_reference(results.records, app.reference)

    search_terms = SearchTerms(case_type="export", manufacture_country=yemen)
    results = search_applications(search_terms, exporter_one_fixture_data.ilb_admin_user)

    assert results.total_rows == 0

    search_terms = SearchTerms(case_type="export", manufacture_country=peru_and_yemen)
    results = search_applications(search_terms, exporter_one_fixture_data.ilb_admin_user)

    assert results.total_rows == 1
    check_export_application_case_reference(results.records, app.reference)


def test_export_search_by_pending_firs(exporter_one_fixture_data: ExportFixtureData):
    com_app = Build.com_application(exporter_one_fixture_data)
    search_terms = SearchTerms(case_type="export", pending_firs=YesNoChoices.yes)
    results = search_applications(search_terms, exporter_one_fixture_data.ilb_admin_user)

    assert results.total_rows == 0

    com_app.further_information_requests.create(
        status=FurtherInformationRequest.OPEN, process_type=FurtherInformationRequest.PROCESS_TYPE
    )

    results = search_applications(search_terms, exporter_one_fixture_data.ilb_admin_user)

    assert results.total_rows == 1
    check_export_application_case_reference(results.records, com_app.reference)


def test_export_search_by_pending_update_reqs(exporter_one_fixture_data: ExportFixtureData):
    com_app = Build.com_application(exporter_one_fixture_data)
    search_terms = SearchTerms(case_type="export", pending_update_reqs=YesNoChoices.yes)
    results = search_applications(search_terms, exporter_one_fixture_data.ilb_admin_user)

    assert results.total_rows == 0

    com_app.update_requests.create(status=UpdateRequest.Status.OPEN)

    results = search_applications(search_terms, exporter_one_fixture_data.ilb_admin_user)

    assert results.total_rows == 1
    check_export_application_case_reference(results.records, com_app.reference)


def test_export_returns_in_progress_applications(exporter_one_fixture_data: ExportFixtureData):
    Build.gmp_application(exporter_one_fixture_data, submit=False)
    submitted_gmp_app = Build.gmp_application(exporter_one_fixture_data)

    search_terms = SearchTerms(case_type="export")
    results = search_applications(search_terms, exporter_one_fixture_data.ilb_admin_user)

    assert results.total_rows == 2

    # The in progress application doesn't have a case reference.
    check_export_application_case_reference(
        results.records, submitted_gmp_app.reference, "Not Assigned"
    )
    assert results.records[1].status == "In Progress"

    search_terms = SearchTerms(case_type="export", case_status=ImpExpStatus.IN_PROGRESS)
    results = search_applications(search_terms, exporter_one_fixture_data.ilb_admin_user)

    assert results.total_rows == 1
    check_export_application_case_reference(results.records, "Not Assigned")
    assert results.records[0].status == "In Progress"


@pytest.mark.parametrize(
    ["case_ref_pattern", "should_match"],
    [
        ("wood/foo/0001", True),
        ("wood%", True),
        ("wood%0001", True),
        ("wood%oo%000%", True),
        ("WOOD%", True),
        ("WOOD%0001", True),
        ("WOOD%OO%000%", True),
        ("%foo/0001", True),
        ("%wood", False),
        ("%WOOD", False),
        ("foo/0001%", False),
    ],
)
def test_wildcard_search(importer_one_fixture_data, case_ref_pattern, should_match):
    """Use case_ref to test the wildcard pattern matching - we don't need to test each field."""

    # Create another record just to ensure its not being returned by mistake
    not_match_app = Build.fa_dfl_application("fa-dfl-app-reference", importer_one_fixture_data)
    wood_app = Build.wood_application("wood-applicant-reference", importer_one_fixture_data)

    # We override the reference so we can test the wildcard matching
    wood_app.reference = "wood/foo/0001"
    wood_app.save()

    not_match_app.reference = "dfl/bar/1110"
    not_match_app.save()

    search_terms = SearchTerms(case_type="import", case_ref=case_ref_pattern)
    results = search_applications(search_terms, importer_one_fixture_data.ilb_admin_user)

    if should_match:
        assert results.total_rows == 1
        assert results.records[0].case_status.case_reference == wood_app.get_reference()

    else:
        assert results.total_rows == 0


def test_case_reference_wildcard_any(
    importer_one_fixture_data: FixtureData,
    exporter_one_fixture_data: ExportFixtureData,
):
    # Import applications
    Build.fa_dfl_application("fa-dfl-app-reference", importer_one_fixture_data)
    Build.wood_application("wood-applicant-reference", importer_one_fixture_data)

    # Export application
    gmp_app = Build.gmp_application(exporter_one_fixture_data)

    search_terms = SearchTerms(case_type="import", case_ref="%")
    results = search_applications(search_terms, exporter_one_fixture_data.ilb_admin_user)

    assert results.total_rows == 2
    check_application_references(
        results.records, "wood-applicant-reference", "fa-dfl-app-reference"
    )

    search_terms = SearchTerms(case_type="export", case_ref="%")
    results = search_applications(search_terms, exporter_one_fixture_data.ilb_admin_user)

    assert results.total_rows == 1
    check_export_application_case_reference(results.records, gmp_app.reference)


def test_search_by_application_contact(
    importer_one_fixture_data: FixtureData, exporter_one_fixture_data: ExportFixtureData
):
    Build.wood_application("wood-applicant-reference", importer_one_fixture_data)
    gmp_app = Build.gmp_application(exporter_one_fixture_data)

    name_search = f"{importer_one_fixture_data.importer_user.first_name}%"

    search_terms = SearchTerms(case_type="import", application_contact=name_search)
    results = search_applications(search_terms, exporter_one_fixture_data.ilb_admin_user)

    assert results.total_rows == 1
    check_application_references(results.records, "wood-applicant-reference")

    search_terms.application_contact = "Not valid"
    results = search_applications(search_terms, exporter_one_fixture_data.ilb_admin_user)
    assert results.total_rows == 0

    name_search = f"%{exporter_one_fixture_data.exporter_user.last_name}"
    search_terms = SearchTerms(case_type="export", application_contact=name_search)
    results = search_applications(search_terms, exporter_one_fixture_data.ilb_admin_user)

    assert results.total_rows == 1
    check_export_application_case_reference(results.records, gmp_app.reference)


def test_import_search_by_licence_type(importer_one_fixture_data: FixtureData):
    wood = Build.wood_application("wood-app-ref", importer_one_fixture_data)

    licence = document_pack.pack_draft_get(wood)
    licence.issue_paper_licence_only = True
    licence.save()

    cdr = document_pack.doc_ref_licence_create(licence, "licence_reference")

    search_terms = SearchTerms(case_type="import", licence_type="electronic")
    results = search_applications(search_terms, importer_one_fixture_data.ilb_admin_user)

    assert results.total_rows == 0

    search_terms = SearchTerms(case_type="import", licence_type="paper")
    results = search_applications(search_terms, importer_one_fixture_data.ilb_admin_user)

    assert results.total_rows == 1
    assert results.records[0].case_status.licence_reference == ""
    assert results.records[0].case_status.licence_reference_link == "#"

    # Only a complete application will show the references
    wood.status = ImpExpStatus.COMPLETED
    wood.save()
    search_terms = SearchTerms(case_type="import", licence_type="paper")
    results = search_applications(search_terms, importer_one_fixture_data.ilb_admin_user)
    assert results.total_rows == 1

    check_application_references(results.records, "wood-app-ref")
    assert results.records[0].case_status.licence_reference == "licence_reference (Paper)"
    assert results.records[0].case_status.licence_reference_link == reverse(
        "case:view-case-document",
        kwargs={
            "application_pk": wood.id,
            "case_type": "import",
            "object_pk": licence.id,
            "casedocumentreference_pk": cdr.id,
        },
    )


def test_import_search_by_chief_usage_status(importer_one_fixture_data: FixtureData):
    wood = Build.wood_application("wood-app-ref", importer_one_fixture_data)

    for value, label in ImportApplication.ChiefUsageTypes.choices:
        wood.chief_usage_status = value
        wood.save()

        search_terms = SearchTerms(case_type="import", chief_usage_status=value)
        results = search_applications(search_terms, importer_one_fixture_data.ilb_admin_user)

        assert results.total_rows == 1
        check_application_references(results.records, "wood-app-ref")


def test_import_country_searches(importer_one_fixture_data: FixtureData):
    Build.fa_sil_application("fa-sil-app", importer_one_fixture_data)

    origin_country = Country.objects.filter(name="Argentina")
    origin_country_multiple = Country.objects.filter(name__in=["Argentina", "Aruba"])

    consignment_country = Country.objects.filter(name="Azerbaijan")
    consignment_country_multiple = Country.objects.filter(name__in=["Azerbaijan", "Aruba"])

    search_terms = SearchTerms(case_type="import", origin_country=origin_country)
    results = search_applications(search_terms, importer_one_fixture_data.ilb_admin_user)
    assert results.total_rows == 1
    check_application_references(results.records, "fa-sil-app")

    search_terms = SearchTerms(case_type="import", origin_country=origin_country_multiple)
    results = search_applications(search_terms, importer_one_fixture_data.ilb_admin_user)
    assert results.total_rows == 1
    check_application_references(results.records, "fa-sil-app")

    search_terms = SearchTerms(case_type="import", consignment_country=consignment_country)
    results = search_applications(search_terms, importer_one_fixture_data.ilb_admin_user)
    assert results.total_rows == 1
    check_application_references(results.records, "fa-sil-app")

    search_terms = SearchTerms(case_type="import", consignment_country=consignment_country_multiple)
    results = search_applications(search_terms, importer_one_fixture_data.ilb_admin_user)
    assert results.total_rows == 1
    check_application_references(results.records, "fa-sil-app")

    search_terms = SearchTerms(case_type="import", consignment_country=origin_country)
    results = search_applications(search_terms, importer_one_fixture_data.ilb_admin_user)
    assert results.total_rows == 0

    search_terms = SearchTerms(case_type="import", origin_country=consignment_country)
    results = search_applications(search_terms, importer_one_fixture_data.ilb_admin_user)
    assert results.total_rows == 0


def test_import_search_by_shipping_year(importer_one_fixture_data: FixtureData):
    # Doesn't have a shipping year
    Build.fa_dfl_application("fa_dfl-app-ref", importer_one_fixture_data)

    # Does have a shipping year
    Build.ironsteel_application("iron-app-ref", importer_one_fixture_data, shipping_year=2022)
    Build.textiles_application("textiles-app-ref", importer_one_fixture_data, shipping_year=2022)
    Build.wood_application("wood-app-ref", importer_one_fixture_data, shipping_year=2022)
    Build.textiles_application(
        "textiles_two-app-ref", importer_one_fixture_data, shipping_year=2023
    )
    Build.wood_application("wood_two-app-ref", importer_one_fixture_data, shipping_year=2024)

    search_terms = SearchTerms(case_type="import", shipping_year="2022")
    results = search_applications(search_terms, importer_one_fixture_data.ilb_admin_user)
    assert results.total_rows == 3
    check_application_references(
        results.records, "wood-app-ref", "textiles-app-ref", "iron-app-ref"
    )

    search_terms.shipping_year = "2023"
    results = search_applications(search_terms, importer_one_fixture_data.ilb_admin_user)
    assert results.total_rows == 1
    check_application_references(results.records, "textiles_two-app-ref")

    search_terms.shipping_year = "2024"
    results = search_applications(search_terms, importer_one_fixture_data.ilb_admin_user)
    assert results.total_rows == 1
    check_application_references(results.records, "wood_two-app-ref")


def test_import_search_by_goods_category(importer_one_fixture_data: FixtureData):
    Build.ironsteel_application(
        "iron-app-ref", importer_one_fixture_data, category_commodity_group="test-1"
    )
    Build.textiles_application(
        "textiles-app-ref", importer_one_fixture_data, category_commodity_group="test-2"
    )

    cp_category = CP_CATEGORIES[3]
    create_test_commodity_group(cp_category, create_test_commodity("test-code"))
    Build.opt_application("opt-app-ref", importer_one_fixture_data, cp_category=cp_category)

    test_pairs = [
        ("test-1", "iron-app-ref"),
        ("test-2", "textiles-app-ref"),
        (cp_category, "opt-app-ref"),
    ]

    for group_code, app_ref in test_pairs:
        group = CommodityGroup.objects.get(group_code=group_code)

        search_terms = SearchTerms(case_type="import", goods_category=group)
        results = search_applications(search_terms, importer_one_fixture_data.ilb_admin_user)
        assert results.total_rows == 1
        check_application_references(results.records, app_ref)

    Build.fa_dfl_application(
        "fa-dfl-ref", importer_one_fixture_data, commodity_code=FirearmCommodity.EX_CHAPTER_93
    )
    Build.fa_sil_application(
        "fa-sil-ref", importer_one_fixture_data, commodity_code=FirearmCommodity.EX_CHAPTER_93
    )
    Build.fa_oil_application(
        "fa-oil-ref", importer_one_fixture_data, commodity_code=FirearmCommodity.EX_CHAPTER_97
    )

    chapter_93 = CommodityGroup.objects.get(group_name=FirearmCommodity.EX_CHAPTER_93.label)
    chapter_97 = CommodityGroup.objects.get(group_name=FirearmCommodity.EX_CHAPTER_97.label)

    search_terms = SearchTerms(case_type="import", goods_category=chapter_93)
    results = search_applications(search_terms, importer_one_fixture_data.ilb_admin_user)

    assert results.total_rows == 2
    check_application_references(results.records, "fa-sil-ref", "fa-dfl-ref")

    search_terms.goods_category = chapter_97
    results = search_applications(search_terms, importer_one_fixture_data.ilb_admin_user)

    assert results.total_rows == 1
    check_application_references(results.records, "fa-oil-ref")

    search_terms.goods_category = None
    results = search_applications(search_terms, importer_one_fixture_data.ilb_admin_user)
    assert results.total_rows == 6


def test_import_search_by_commodity_code(importer_one_fixture_data: FixtureData):
    Build.derogation_application(
        "derogation-app", importer_one_fixture_data, commodity_code="xx111111xx"
    )
    Build.ironsteel_application(
        "ironsteel-app", importer_one_fixture_data, commodity_code="xx222222xx"
    )
    Build.opt_application(
        "opt-app",
        importer_one_fixture_data,
        cp_commodity_codes=["xx333333xx"],
        teg_commodity_codes=["xx444444xx"],
    )
    Build.sanctionadhoc_application(
        "sanctionadhoc-app", importer_one_fixture_data, commodity_codes=["xx555555xx"]
    )
    Build.sps_application("sps-app", importer_one_fixture_data, commodity_code="xx666666xx")
    Build.textiles_application(
        "textiles-app", importer_one_fixture_data, commodity_code="xx777777xx"
    )
    Build.wood_application("wood-app", importer_one_fixture_data, commodity_code="xx888888xx")

    # Check single wildcard returns all apps
    search_terms = SearchTerms(case_type="import", commodity_code="%")
    results = search_applications(search_terms, importer_one_fixture_data.ilb_admin_user)
    assert results.total_rows == 7

    # Search for one record by app type
    search_terms = SearchTerms(
        case_type="import",
        commodity_code="xx111111xx",
        app_type=ImportApplicationType.Types.DEROGATION,
    )
    results = search_applications(search_terms, importer_one_fixture_data.ilb_admin_user)
    assert results.total_rows == 1
    check_application_references(results.records, "derogation-app")

    # Search for matching record using commodity code:
    search_pairs = [
        ("xx111111xx", "derogation-app"),
        ("xx1%1xx", "derogation-app"),
        ("xx222222xx", "ironsteel-app"),
        ("xx2%2xx", "ironsteel-app"),
        ("xx333333xx", "opt-app"),
        ("xx3%3xx", "opt-app"),
        ("xx444444xx", "opt-app"),
        ("xx4%4xx", "opt-app"),
        ("xx555555xx", "sanctionadhoc-app"),
        ("xx5%5xx", "sanctionadhoc-app"),
        ("xx666666xx", "sps-app"),
        ("xx6%6xx", "sps-app"),
        ("xx777777xx", "textiles-app"),
        ("xx7%7xx", "textiles-app"),
        ("xx888888xx", "wood-app"),
        ("xx8%8xx", "wood-app"),
    ]

    for search_term, app_ref in search_pairs:
        search_terms = SearchTerms(case_type="import", commodity_code=search_term)
        results = search_applications(search_terms, importer_one_fixture_data.ilb_admin_user)

        assert results.total_rows == 1
        check_application_references(results.records, app_ref)


def test_import_search_by_document_reference(importer_one_fixture_data: FixtureData):
    wood = Build.wood_application("wood-app-1-ref", importer_one_fixture_data)
    licence = document_pack.pack_draft_get(wood)
    licence.status = DocumentPackBase.Status.ACTIVE
    licence.save()
    document_pack.doc_ref_licence_create(licence, "GBSIL0000001B")

    wood_2 = Build.wood_application("wood-app-2-ref", importer_one_fixture_data)
    licence = document_pack.pack_draft_get(wood_2)
    licence.status = DocumentPackBase.Status.ACTIVE
    licence.save()
    document_pack.doc_ref_licence_create(licence, "GBSIL0000002C")

    # Search exact
    search_terms = SearchTerms(case_type="import", licence_ref="GBSIL0000001B")
    results = search_applications(search_terms, importer_one_fixture_data.ilb_admin_user)
    assert results.total_rows == 1
    check_application_references(results.records, "wood-app-1-ref")

    # Search wildcard matches one app
    search_terms = SearchTerms(case_type="import", licence_ref="%2c")
    results = search_applications(search_terms, importer_one_fixture_data.ilb_admin_user)
    assert results.total_rows == 1
    check_application_references(results.records, "wood-app-2-ref")

    # Search wildcard matches both
    search_terms = SearchTerms(case_type="import", licence_ref="GB%0000%")
    results = search_applications(search_terms, importer_one_fixture_data.ilb_admin_user)
    assert results.total_rows == 2
    check_application_references(results.records, "wood-app-2-ref", "wood-app-1-ref")

    # Search filters excludes archived licences.
    licence = wood.licences.first()
    licence.status = DocumentPackBase.Status.ARCHIVED
    licence.save()

    search_terms = SearchTerms(case_type="import", licence_ref="GBSIL0000001B")
    results = search_applications(search_terms, importer_one_fixture_data.ilb_admin_user)
    assert results.total_rows == 0


def test_export_search_by_document_reference(exporter_one_fixture_data: ExportFixtureData):
    cfs = Build.cfs_application(exporter_one_fixture_data)
    cert = document_pack.pack_draft_get(cfs)
    cert.status = DocumentPackBase.Status.ACTIVE
    cert.save()
    pairs = (
        ("CFS/2021/00001", "Finland"),
        ("CFS/2021/00002", "Germany"),
        ("CFS/2021/00003", "Poland"),
    )
    for ref, country in pairs:
        document_pack.doc_ref_certificate_create(
            cert, ref, country=Country.objects.get(name=country)
        )

    # Search exact
    search_terms = SearchTerms(case_type="export", licence_ref="CFS/2021/00002")
    results = search_applications(search_terms, exporter_one_fixture_data.ilb_admin_user)
    assert results.total_rows == 1
    check_export_application_case_reference(results.records, cfs.reference)

    # Search wildcard matches one app
    search_terms = SearchTerms(case_type="export", licence_ref="%/2021/%")
    results = search_applications(search_terms, exporter_one_fixture_data.ilb_admin_user)
    assert results.total_rows == 1
    check_export_application_case_reference(results.records, cfs.reference)

    # Search filters excludes archived licences.
    cert.status = DocumentPackBase.Status.ARCHIVED
    cert.save()

    search_terms = SearchTerms(case_type="export", licence_ref="CFS/2021/0000%")
    results = search_applications(search_terms, exporter_one_fixture_data.ilb_admin_user)
    assert results.total_rows == 0


def test_reassignment_search(
    importer_one_fixture_data: FixtureData,
    exporter_one_fixture_data: ExportFixtureData,
    client,
    ilb_admin_user,
):
    wood_app = Build.wood_application("wood-app-1", importer_one_fixture_data)
    textiles_app = Build.textiles_application("textiles-app-1", importer_one_fixture_data)

    # We need to be the icms case officer to post to the take-ownership endpoint
    client.force_login(ilb_admin_user)

    assert wood_app.status == ImpExpStatus.SUBMITTED

    take_ownership_url = reverse(
        "case:take-ownership", kwargs={"application_pk": wood_app.pk, "case_type": "import"}
    )
    response = client.post(take_ownership_url)
    assert response.status_code == 302

    take_ownership_url = reverse(
        "case:take-ownership", kwargs={"application_pk": textiles_app.pk, "case_type": "import"}
    )
    response = client.post(take_ownership_url)
    assert response.status_code == 302

    wood_app.refresh_from_db()
    assert wood_app.status == ImpExpStatus.PROCESSING
    assert wood_app.case_owner == ilb_admin_user

    search_terms = SearchTerms(
        case_type="import",
        reassignment_search=True,
    )
    results = search_applications(search_terms, importer_one_fixture_data.ilb_admin_user)

    assert results.total_rows == 2
    check_application_references(results.records, "textiles-app-1", "wood-app-1")

    # Override the case owner to test "reassignment_user"
    textiles_app.case_owner = importer_one_fixture_data.importer_user
    textiles_app.save()

    search_terms = SearchTerms(
        case_type="import", reassignment_search=True, reassignment_user=ilb_admin_user
    )
    results = search_applications(search_terms, importer_one_fixture_data.ilb_admin_user)

    assert results.total_rows == 1
    check_application_references(results.records, "wood-app-1")

    # Test Export applications work with reassignment searching
    gmp_app = Build.gmp_application(exporter_one_fixture_data)
    search_terms = SearchTerms(
        case_type="export", reassignment_search=True, reassignment_user=ilb_admin_user
    )
    results = search_applications(search_terms, exporter_one_fixture_data.ilb_admin_user)

    assert results.total_rows == 0
    take_ownership_url = reverse(
        "case:take-ownership", kwargs={"application_pk": gmp_app.pk, "case_type": "export"}
    )
    response = client.post(take_ownership_url)
    assert response.status_code == 302

    results = search_applications(search_terms, exporter_one_fixture_data.ilb_admin_user)

    assert results.total_rows == 1
    check_export_application_case_reference(results.records, gmp_app.reference)


def test_can_search_refused_application(fa_dfl_app_submitted, ilb_admin_client, ilb_admin_user):
    app = fa_dfl_app_submitted
    # Take ownership
    r = ilb_admin_client.post(CaseURLS.take_ownership(app.pk))
    assert r.status_code == 302

    # Create a checklist
    DFLChecklist.objects.create(
        import_application=app,
        deactivation_certificate_attached=YesNoChoices.yes,
        deactivation_certificate_issued=YesNoChoices.yes,
        case_update=YesNoChoices.yes,
        fir_required=YesNoChoices.yes,
        response_preparation=True,
        validity_period_correct=YesNoChoices.yes,
        endorsements_listed=YesNoChoices.yes,
        authorisation=True,
    )

    # Refuse the application
    app.decision = app.REFUSE
    app.save()

    r = ilb_admin_client.post(CaseURLS.start_authorisation(app.pk))
    assert r.status_code == 302

    # Test we can search by it
    search_terms = SearchTerms(case_type="import", case_status=ImpExpStatus.COMPLETED)
    results = search_applications(search_terms, ilb_admin_user)
    assert results.total_rows == 1
    assert len(results.records) == 1


class TestImporterSearchPermissions:
    @pytest.fixture(autouse=True)
    def setup(
        self, db, importer_one_fixture_data: FixtureData, importer_two_fixture_data: FixtureData
    ):
        self.importer_one_conf = importer_one_fixture_data
        self.importer_two_conf = importer_two_fixture_data

    def test_importer_user_only_sees_correct_applications(self):
        Build.fa_dfl_application("fa-dfl-ref-importer-1-app-1", self.importer_one_conf)
        Build.fa_dfl_application("fa-dfl-ref-importer-2-app-1", self.importer_two_conf)

        # Check importer one user only sees records linked to importer one
        search_terms = SearchTerms(case_type="import")

        results = search_applications(search_terms, self.importer_one_conf.importer_user)
        assert results.total_rows == 1
        assert len(results.records) == 1
        check_application_references(results.records, "fa-dfl-ref-importer-1-app-1")

        # Check importer two user only sees records linked to importer two
        results = search_applications(search_terms, self.importer_two_conf.importer_user)
        assert results.total_rows == 1
        assert len(results.records) == 1

        check_application_references(results.records, "fa-dfl-ref-importer-2-app-1")

    def test_importer_user_sees_linked_agent_applications(self):
        Build.fa_dfl_application("fa-dfl-ref-importer-1-app-1", self.importer_one_conf)
        Build.fa_dfl_application("fa-dfl-ref-importer-1-app-2", self.importer_one_conf)
        Build.fa_dfl_application(
            "fa-dfl-ref-importer-1-app-3-agent-app", self.importer_one_conf, agent_application=True
        )
        Build.fa_dfl_application("fa-dfl-ref-importer-2-app-1", self.importer_two_conf)

        search_terms = SearchTerms(case_type="import")
        results = search_applications(search_terms, self.importer_one_conf.importer_user)
        assert results.total_rows == 3
        assert len(results.records) == 3

        check_application_references(
            results.records,
            "fa-dfl-ref-importer-1-app-1",
            "fa-dfl-ref-importer-1-app-2",
            "fa-dfl-ref-importer-1-app-3-agent-app",
            sort_results=True,
        )

    def test_agent_only_sees_correct_applications(self):
        Build.fa_dfl_application("fa-dfl-ref-importer-1", self.importer_one_conf)
        Build.fa_dfl_application(
            "fa-dfl-ref-importer-1-agent-app", self.importer_one_conf, agent_application=True
        )

        search_terms = SearchTerms(case_type="import")
        results = search_applications(search_terms, self.importer_one_conf.agent_user)
        assert results.total_rows == 1
        assert len(results.records) == 1

        check_application_references(results.records, "fa-dfl-ref-importer-1-agent-app")


class TestExporterSearchPermissions:
    @pytest.fixture(autouse=True)
    def setup(
        self,
        db,
        exporter_one_fixture_data: ExportFixtureData,
        exporter_two_fixture_data: ExportFixtureData,
    ):
        self.exporter_one_conf = exporter_one_fixture_data
        self.exporter_two_conf = exporter_two_fixture_data

    def test_export_permissions(self):
        exp_1_app_1 = Build.com_application(self.exporter_one_conf)
        exp_1_app_2 = Build.com_application(self.exporter_one_conf)
        exp_1_app_3_agent_app = Build.com_application(
            self.exporter_one_conf, agent_application=True
        )

        exp_2_app_1 = Build.com_application(self.exporter_two_conf)
        exp_2_app_2 = Build.com_application(self.exporter_two_conf)

        #
        # Test admin can see all applications
        #
        search_terms = SearchTerms(case_type="export")
        results = search_applications(search_terms, self.exporter_one_conf.ilb_admin_user)
        assert results.total_rows == 5
        assert len(results.records) == 5

        #
        # Test exporter one contact can see all "Test Exporter One" applications
        #
        search_terms = SearchTerms(case_type="export")
        results = search_applications(search_terms, self.exporter_one_conf.exporter_user)
        assert results.total_rows == 3
        assert len(results.records) == 3
        check_export_application_case_reference(
            results.records,
            exp_1_app_1.reference,
            exp_1_app_2.reference,
            exp_1_app_3_agent_app.reference,
            sort_results=True,
        )

        #
        # Test exporter one agent can only see "Test Exporter 1 Agent 1" applications
        #
        search_terms = SearchTerms(case_type="export")
        results = search_applications(search_terms, self.exporter_one_conf.exporter_agent_user)
        assert results.total_rows == 1
        assert len(results.records) == 1
        check_export_application_case_reference(results.records, exp_1_app_3_agent_app.reference)

        #
        # Test exporter two contact can see all "Test Exporter 2" applications
        #
        search_terms = SearchTerms(case_type="export")
        results = search_applications(search_terms, self.exporter_two_conf.exporter_user)
        assert results.total_rows == 2
        assert len(results.records) == 2
        check_export_application_case_reference(
            results.records, exp_2_app_1.reference, exp_2_app_2.reference, sort_results=True
        )


def check_application_references(
    applications: list[types.ResultRow], *references, sort_results=False
):
    """Check the returned applications match the supplied references

    Sort results if we don't care about the order
    """

    expected = sorted(references) if sort_results else list(references)

    actual = (app.case_status.applicant_reference for app in applications)
    actual = sorted(actual) if sort_results else list(actual)  # type: ignore[type-var, assignment]

    assert expected == actual


def check_export_application_case_reference(
    applications: list[types.ResultRow], *app_references, sort_results=False
):
    expected = sorted(app_references) if sort_results else list(app_references)

    actual = (app.case_reference for app in applications)
    actual = sorted(actual) if sort_results else list(actual)  # type: ignore[type-var, assignment]

    assert expected == actual


def check_commodity_details(
    actual_details: types.CommodityDetails,
    *,
    expected_origin_country: str | None = None,
    expected_consignment_country: str | None = None,
    expected_goods_category: str | None = None,
    expected_shipping_year: int | None = None,
    expected_commodity_codes: list[str] | None = None,
):
    assert expected_origin_country == actual_details.origin_country
    assert expected_consignment_country == actual_details.consignment_country
    assert expected_goods_category == actual_details.goods_category
    assert expected_shipping_year == actual_details.shipping_year
    assert expected_commodity_codes == actual_details.commodity_codes
