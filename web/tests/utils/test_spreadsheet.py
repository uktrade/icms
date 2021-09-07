import io

from openpyxl import load_workbook

from web.utils.spreadsheet import XlsxConfig, generate_xlsx_file


def test_generate_xlsx_spreadsheet():
    config = XlsxConfig()
    test_data = [
        ["H1", "H2", "H3"],
        ["R1C1", "R1C2", "R1C3"],
        ["R2C1", "R2C2", "R2C3"],
        ["R3C1", "R3C2", "R3C3"],
    ]
    config.header.data = test_data[0]
    config.rows = test_data[1:]
    config.sheet_name = "My Sheet"
    xlsx_data = generate_xlsx_file(config)

    workbook = load_workbook(filename=io.BytesIO(xlsx_data))

    assert workbook.sheetnames == ["My Sheet"]

    sheet = workbook["My Sheet"]

    for row, row_data in enumerate(sheet.values):
        assert list(row_data) == test_data[row]


def test_generate_xlsx_template():
    config = XlsxConfig()
    header_data = ["H1", "H2", "H3"]
    config.header.data = header_data
    xlsx_data = generate_xlsx_file(config)

    workbook = load_workbook(filename=io.BytesIO(xlsx_data))
    sheet = workbook["Sheet1"]
    cols = sheet.max_column
    rows = sheet.max_row
    header = next(sheet.values)

    assert cols == 3
    assert rows == 1
    assert list(header) == header_data
