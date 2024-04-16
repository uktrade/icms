import io
from http import HTTPStatus
from unittest import mock

import pytest
from django.urls import reverse, reverse_lazy
from openpyxl import load_workbook
from pytest_django.asserts import assertTemplateUsed

from web.domains.cat.forms import CATFilter
from web.domains.cat.views import CatSteps
from web.domains.country.models import Country
from web.models import (
    CertificateApplicationTemplate,
    CertificateOfFreeSaleApplicationTemplate,
    CertificateOfGoodManufacturingPracticeApplicationTemplate,
    ExportApplicationType,
    ProductLegislation,
)
from web.models.shared import AddressEntryType, YesNoChoices
from web.tests.auth import AuthTestCase
from web.tests.domains.case.export.test_utils import (
    create_dummy_config,
    create_dummy_xlsx_file,
)


class TestCATListView(AuthTestCase):
    url = reverse_lazy("cat:list")

    def test_template_context(self):
        response = self.exporter_client.get(self.url)

        assert response.status_code == HTTPStatus.OK
        assertTemplateUsed(response, "web/domains/cat/list.html")

        assert isinstance(response.context["filter"], CATFilter)

    def test_filter_queryset_by_name(self):
        for name in ("Foo", "Bar", "Baz"):
            CertificateApplicationTemplate.objects.create(
                owner=self.exporter_user,
                name=name,
                application_type=ExportApplicationType.Types.GMP,
            )

        # Filtering with query parameters.
        response = self.exporter_client.get(self.url, {"name": "foo"})

        assert response.status_code == HTTPStatus.OK
        assert [t.name for t in response.context["templates"]] == ["Foo"]

    def test_filter_defaults_to_current_templates(self):
        foo = CertificateApplicationTemplate.objects.create(
            owner=self.exporter_user,
            name="foo",
            application_type=ExportApplicationType.Types.GMP,
            is_active=True,
        )
        CertificateApplicationTemplate.objects.create(
            owner=self.exporter_user,
            name="bar",
            application_type=ExportApplicationType.Types.GMP,
            is_active=False,
        )

        response = self.exporter_client.get(self.url, {})

        assert response.status_code == HTTPStatus.OK
        # The archived template is not shown.
        assert [t.pk for t in response.context["templates"]] == [foo.pk]

    def test_filter_is_active(self):
        active_t = CertificateApplicationTemplate.objects.create(
            owner=self.exporter_user,
            name="active template",
            application_type=ExportApplicationType.Types.GMP,
        )
        inactive_t = CertificateApplicationTemplate.objects.create(
            owner=self.exporter_user,
            name="inactive template",
            application_type=ExportApplicationType.Types.GMP,
            is_active=False,
        )

        response = self.exporter_client.get(self.url, data={"is_active": "True"})
        assert response.status_code == HTTPStatus.OK
        assert [t.pk for t in response.context["templates"]] == [active_t.pk]

        response = self.exporter_client.get(self.url, data={"is_active": "False"})
        assert response.status_code == HTTPStatus.OK
        assert [t.pk for t in response.context["templates"]] == [inactive_t.pk]

        response = self.exporter_client.get(self.url, data={"is_active": ""})
        assert response.status_code == HTTPStatus.OK
        assert [t.pk for t in response.context["templates"]] == [inactive_t.pk, active_t.pk]

    def test_permission_filter(self):
        exporter_one_template = CertificateApplicationTemplate.objects.create(
            owner=self.exporter_user,
            name="active template",
            application_type=ExportApplicationType.Types.GMP,
        )
        exporter_two_template = CertificateApplicationTemplate.objects.create(
            owner=self.exporter_two_user,
            name="inactive template",
            application_type=ExportApplicationType.Types.GMP,
            is_active=False,
        )

        response = self.exporter_client.get(self.url, data={"is_active": ""})
        assert response.status_code == HTTPStatus.OK
        assert [t.pk for t in response.context["templates"]] == [exporter_one_template.pk]

        response = self.exporter_two_client.get(self.url, data={"is_active": ""})
        assert response.status_code == HTTPStatus.OK
        assert [t.pk for t in response.context["templates"]] == [exporter_two_template.pk]


class TestCATCreateView(AuthTestCase):
    @pytest.mark.parametrize(
        ["app_type", "related_model"],
        [
            (ExportApplicationType.Types.FREE_SALE, "cfs_template"),
            (ExportApplicationType.Types.MANUFACTURE, "com_template"),
            (ExportApplicationType.Types.GMP, "gmp_template"),
        ],
    )
    def test_can_create_template(self, app_type, related_model):
        with pytest.raises(CertificateApplicationTemplate.DoesNotExist):
            CertificateApplicationTemplate.objects.get()

        url = reverse("cat:create")
        data = {
            "name": f"{app_type} Template name",
            "description": "Template description",
            "application_type": app_type,
            "sharing": CertificateApplicationTemplate.SharingStatuses.PRIVATE,
        }
        if app_type == ExportApplicationType.Types.FREE_SALE:
            data["template_country"] = CertificateApplicationTemplate.CountryType.GB

        response = self.exporter_client.post(url, data)

        assert response.status_code == HTTPStatus.FOUND
        assert response["Location"].startswith("/cat/edit/")

        template = CertificateApplicationTemplate.objects.get(name=f"{app_type} Template name")
        assert template.name == f"{app_type} Template name"
        related_template = getattr(template, related_model)
        assert related_template is not None

        if app_type == ExportApplicationType.Types.FREE_SALE:
            assert related_template.schedules.count() == 1
            assert not template.is_ni_template


class TestCATEditView(AuthTestCase):
    def test_permission_denied(self, cfs_cat: CertificateApplicationTemplate):
        url = reverse("cat:edit", kwargs={"cat_pk": cfs_cat.pk})
        response = self.importer_client.get(url)

        assert response.status_code == HTTPStatus.FORBIDDEN

    def test_edit_cfs_template(self, cfs_cat: CertificateApplicationTemplate):
        cat = cfs_cat
        # Test template edit view
        url = reverse("cat:edit", kwargs={"cat_pk": cat.pk})
        response = self.exporter_client.get(url)

        assert response.status_code == HTTPStatus.OK
        assertTemplateUsed(response, "web/domains/cat/edit.html")

        form_data = {
            "name": cat.name,
            "description": "Updated description",
            "sharing": CertificateApplicationTemplate.SharingStatuses.PRIVATE,
            "template_country": CertificateApplicationTemplate.CountryType.GB,
        }
        response = self.exporter_client.post(url, data=form_data)

        assert response.status_code == HTTPStatus.FOUND
        cat.refresh_from_db()
        assert cat.description == "Updated description"

        # Test editing CFS step form
        url = reverse("cat:edit-step", kwargs={"cat_pk": cat.pk, "step": CatSteps.CFS})
        response = self.exporter_client.get(url)
        assert response.status_code == HTTPStatus.OK
        assertTemplateUsed(response, "web/domains/cat/cfs/edit.html")
        valid_countries = ExportApplicationType.objects.get(
            type_code=ExportApplicationType.Types.FREE_SALE
        ).country_group.countries.filter(is_active=True)
        first_vc = valid_countries.first()
        form_data = {"countries": [first_vc.pk]}
        response = self.exporter_client.post(url, data=form_data)

        assert response.status_code == HTTPStatus.FOUND
        cat.refresh_from_db()
        assert cat.cfs_template.countries.first() == first_vc

        # Test editing CFS Schedule step form
        url = reverse(
            "cat:edit-step-related",
            kwargs={
                "cat_pk": cat.pk,
                "step": CatSteps.CFS_SCHEDULE,
                "step_pk": cat.cfs_template.schedules.first().pk,
            },
        )
        response = self.exporter_client.get(url)
        assert response.status_code == HTTPStatus.OK
        assertTemplateUsed(response, "web/domains/cat/cfs/edit-schedule.html")

        cfs_com = Country.objects.first()
        form_data = {
            "exporter_status": "MANUFACTURER",
            "brand_name_holder": YesNoChoices.yes,
            "product_eligibility": "SOLD_ON_UK_MARKET",
            "goods_placed_on_uk_market": YesNoChoices.yes,
            "goods_export_only": YesNoChoices.yes,
            "any_raw_materials": YesNoChoices.yes,
            "final_product_end_use": "End Use or Final Product value",
            "country_of_manufacture": cfs_com.pk,
            "schedule_statements_accordance_with_standards": True,
            "schedule_statements_is_responsible_person": False,
        }
        response = self.exporter_client.post(url, data=form_data)

        assert response.status_code == HTTPStatus.FOUND
        cat.refresh_from_db()
        schedule = cat.cfs_template.schedules.first()
        assert schedule.exporter_status == "MANUFACTURER"
        assert schedule.brand_name_holder == YesNoChoices.yes
        assert schedule.product_eligibility == "SOLD_ON_UK_MARKET"
        assert schedule.goods_placed_on_uk_market == YesNoChoices.yes
        assert schedule.goods_export_only == YesNoChoices.yes
        assert schedule.any_raw_materials == YesNoChoices.yes
        assert schedule.final_product_end_use == "End Use or Final Product value"
        assert schedule.country_of_manufacture == cfs_com
        assert schedule.schedule_statements_accordance_with_standards is True
        assert schedule.schedule_statements_is_responsible_person is False

    def test_edit_com_template(self, com_cat: CertificateApplicationTemplate):
        cat = com_cat

        # Test template edit view
        url = reverse("cat:edit", kwargs={"cat_pk": cat.pk})
        response = self.exporter_client.get(url)

        assert response.status_code == HTTPStatus.OK
        assertTemplateUsed(response, "web/domains/cat/edit.html")

        form_data = {
            "name": cat.name,
            "description": "Updated description",
            "sharing": CertificateApplicationTemplate.SharingStatuses.PRIVATE,
        }
        response = self.exporter_client.post(url, data=form_data)

        assert response.status_code == HTTPStatus.FOUND
        cat.refresh_from_db()
        assert cat.description == "Updated description"

        # Test editing COM step form
        url = reverse("cat:edit-step", kwargs={"cat_pk": cat.pk, "step": CatSteps.COM})
        response = self.exporter_client.get(url)
        assert response.status_code == HTTPStatus.OK
        assertTemplateUsed(response, "web/domains/cat/edit.html")

        form_data = {
            "is_pesticide_on_free_sale_uk": False,
            "is_manufacturer": True,
            "product_name": "Test product name",
            "chemical_name": "Test chemical name",
            "manufacturing_process": "Test manufacturing process",
        }
        response = self.exporter_client.post(url, data=form_data)
        assert response.status_code == HTTPStatus.FOUND

        assert not cat.com_template.is_pesticide_on_free_sale_uk
        assert cat.com_template.is_manufacturer
        assert cat.com_template.product_name == "Test product name"
        assert cat.com_template.chemical_name == "Test chemical name"
        assert cat.com_template.manufacturing_process == "Test manufacturing process"

    def test_edit_gmp_template(self, gmp_cat: CertificateApplicationTemplate):
        cat = gmp_cat
        # Test template edit view
        url = reverse("cat:edit", kwargs={"cat_pk": cat.pk})
        response = self.exporter_client.get(url)

        assert response.status_code == HTTPStatus.OK
        assertTemplateUsed(response, "web/domains/cat/edit.html")

        form_data = {
            "name": cat.name,
            "description": "Updated description",
            "sharing": CertificateApplicationTemplate.SharingStatuses.PRIVATE,
        }
        response = self.exporter_client.post(url, data=form_data)

        assert response.status_code == HTTPStatus.FOUND
        cat.refresh_from_db()
        assert cat.description == "Updated description"

        # Test editing GMP step form
        url = reverse("cat:edit-step", kwargs={"cat_pk": cat.pk, "step": CatSteps.GMP})
        response = self.exporter_client.get(url)
        assert response.status_code == HTTPStatus.OK
        assertTemplateUsed(response, "web/domains/cat/edit.html")

        form_data = {
            "brand_name": "Test brand name",
            "is_manufacturer": YesNoChoices.yes,
            "is_responsible_person": YesNoChoices.yes,
            "manufacturer_address": "Test manufacturer address",
            "manufacturer_address_entry_type": AddressEntryType.MANUAL,
            "manufacturer_country": "GB",
            "manufacturer_name": "Test manufacturer name",
            "manufacturer_postcode": "S12SS",  # /PS-IGNORE
            "responsible_person_address": "Test responsible person address",
            "responsible_person_address_entry_type": AddressEntryType.MANUAL,
            "responsible_person_country": "GB",
            "responsible_person_name": "Test responsible person name",
            "responsible_person_postcode": "S12SS",  # /PS-IGNORE
        }
        response = self.exporter_client.post(url, data=form_data)
        assert response.status_code == HTTPStatus.FOUND

        assert cat.gmp_template.brand_name == "Test brand name"
        assert cat.gmp_template.is_manufacturer == YesNoChoices.yes
        assert cat.gmp_template.is_responsible_person == YesNoChoices.yes
        assert cat.gmp_template.manufacturer_address == "Test manufacturer address"
        assert cat.gmp_template.manufacturer_address_entry_type == AddressEntryType.MANUAL
        assert cat.gmp_template.manufacturer_country == "GB"
        assert cat.gmp_template.manufacturer_name == "Test manufacturer name"
        assert cat.gmp_template.manufacturer_postcode == "S12SS"  # /PS-IGNORE
        assert cat.gmp_template.responsible_person_address == "Test responsible person address"
        assert cat.gmp_template.responsible_person_address_entry_type == AddressEntryType.MANUAL
        assert cat.gmp_template.responsible_person_country == "GB"
        assert cat.gmp_template.responsible_person_name == "Test responsible person name"
        assert cat.gmp_template.responsible_person_postcode == "S12SS"  # /PS-IGNORE


class TestCATArchiveView(AuthTestCase):
    def test_archive_a_template(self):
        cat = CertificateApplicationTemplate.objects.create(
            owner=self.exporter_user,
            name="GMP template",
            application_type=ExportApplicationType.Types.GMP,
        )
        CertificateOfGoodManufacturingPracticeApplicationTemplate.objects.create(template=cat)
        assert cat.is_active is True

        url = reverse("cat:archive", kwargs={"cat_pk": cat.pk})
        response = self.exporter_client.post(url)

        assert response.status_code == HTTPStatus.FOUND
        assert response["Location"] == "/cat/"

        cat.refresh_from_db()

        assert cat.is_active is False

    def test_permission_denied(self):
        cat = CertificateApplicationTemplate.objects.create(
            owner=self.exporter_user,
            name="CFS template",
            application_type=ExportApplicationType.Types.FREE_SALE,
            template_country=CertificateApplicationTemplate.CountryType.GB,
        )
        CertificateOfFreeSaleApplicationTemplate.objects.create(template=cat)
        url = reverse("cat:archive", kwargs={"cat_pk": cat.pk})

        response = self.importer_client.get(url)

        assert response.status_code == HTTPStatus.FORBIDDEN


class TestCATRestoreView(AuthTestCase):
    def test_restore_a_template(self):
        cat = CertificateApplicationTemplate.objects.create(
            owner=self.exporter_user,
            name="GMP template",
            application_type=ExportApplicationType.Types.GMP,
            is_active=False,
        )
        CertificateOfGoodManufacturingPracticeApplicationTemplate.objects.create(template=cat)
        assert cat.is_active is False

        url = reverse("cat:restore", kwargs={"cat_pk": cat.pk})
        response = self.exporter_client.post(url)

        assert response.status_code == HTTPStatus.FOUND
        assert response["Location"] == "/cat/"

        cat.refresh_from_db()

        assert cat.is_active is True

    def test_permission_denied(self):
        cat = CertificateApplicationTemplate.objects.create(
            owner=self.exporter_user,
            name="CFS template",
            application_type=ExportApplicationType.Types.FREE_SALE,
            template_country=CertificateApplicationTemplate.CountryType.GB,
            is_active=False,
        )
        CertificateOfFreeSaleApplicationTemplate.objects.create(template=cat)
        url = reverse("cat:restore", kwargs={"cat_pk": cat.pk})

        response = self.importer_client.get(url)

        assert response.status_code == HTTPStatus.FORBIDDEN


class TestCATReadOnlyView(AuthTestCase):
    def test_show_disabled_form_for_template(self):
        cat = CertificateApplicationTemplate.objects.create(
            owner=self.exporter_user,
            name="CFS template",
            application_type=ExportApplicationType.Types.FREE_SALE,
            template_country=CertificateApplicationTemplate.CountryType.GB,
        )
        CertificateOfFreeSaleApplicationTemplate.objects.create(template=cat)
        url = reverse("cat:view", kwargs={"cat_pk": cat.pk})

        response = self.exporter_client.get(url)

        assert response.status_code == HTTPStatus.OK
        assert response.context["read_only"] is True

    def test_show_disabled_form_for_step(self):
        cat = CertificateApplicationTemplate.objects.create(
            owner=self.exporter_user,
            name="CFS template",
            application_type=ExportApplicationType.Types.FREE_SALE,
            template_country=CertificateApplicationTemplate.CountryType.GB,
        )
        CertificateOfFreeSaleApplicationTemplate.objects.create(template=cat)
        url = reverse("cat:view-step", kwargs={"cat_pk": cat.pk, "step": "cfs"})

        response = self.exporter_client.get(url)

        assert response.status_code == HTTPStatus.OK
        assert response.context["read_only"] is True


class TestCFSManufacturerUpdateView(AuthTestCase):
    @pytest.fixture(autouse=True)
    def setup(self, _setup, cfs_cat):
        self.app = cfs_cat
        self.url = reverse(
            "cat:cfs-manufacturer-update",
            kwargs={
                "cat_pk": self.app.pk,
                "schedule_template_pk": self.app.cfs_template.schedules.first().pk,
            },
        )

    def test_permission(self):
        response = self.importer_client.get(self.url)
        assert response.status_code == HTTPStatus.FORBIDDEN

        response = self.exporter_two_client.get(self.url)
        assert response.status_code == HTTPStatus.FORBIDDEN

        response = self.exporter_client.get(self.url)
        assert response.status_code == HTTPStatus.OK

    def test_post(self):
        form_data = {
            "manufacturer_name": "test manufacturer name",
            "manufacturer_address_entry_type": AddressEntryType.MANUAL,
            "manufacturer_postcode": "S12SS",  # /PS-IGNORE
            "manufacturer_address": "Test manufacturer address",
        }
        response = self.exporter_client.post(self.url, data=form_data)
        assert response.status_code == HTTPStatus.FOUND

        self.app.refresh_from_db()
        schedule = self.app.cfs_template.schedules.first()
        assert schedule.manufacturer_name == "test manufacturer name"
        assert schedule.manufacturer_address_entry_type == AddressEntryType.MANUAL
        assert schedule.manufacturer_postcode == "S12SS"  # /PS-IGNORE
        assert schedule.manufacturer_address == "Test manufacturer address"


class TestCFSManufacturerDeleteView(AuthTestCase):
    @pytest.fixture(autouse=True)
    def setup(self, _setup, cfs_cat):
        self.app = cfs_cat
        self.url = reverse(
            "cat:cfs-manufacturer-delete",
            kwargs={
                "cat_pk": self.app.pk,
                "schedule_template_pk": self.app.cfs_template.schedules.first().pk,
            },
        )
        schedule = self.app.cfs_template.schedules.first()
        schedule.manufacturer_name = "test manufacturer name"
        schedule.manufacturer_address_entry_type = AddressEntryType.SEARCH
        schedule.manufacturer_postcode = "S12SS"  # /PS-IGNORE
        schedule.manufacturer_address = "Test manufacturer address"
        schedule.save()

    def test_post_only(self):
        response = self.exporter_client.get(self.url)
        assert response.status_code == HTTPStatus.METHOD_NOT_ALLOWED

    def test_permission(self):
        response = self.importer_client.post(self.url)
        assert response.status_code == HTTPStatus.FORBIDDEN

        response = self.exporter_two_client.post(self.url)
        assert response.status_code == HTTPStatus.FORBIDDEN

    def test_post(self):
        response = self.exporter_client.post(self.url)
        assert response.status_code == HTTPStatus.FOUND

        self.app.refresh_from_db()
        schedule = self.app.cfs_template.schedules.first()
        assert schedule.manufacturer_name is None
        assert schedule.manufacturer_address_entry_type == AddressEntryType.MANUAL
        assert schedule.manufacturer_postcode is None
        assert schedule.manufacturer_address is None


class TestCFSScheduleTemplateAddView(AuthTestCase):
    @pytest.fixture(autouse=True)
    def setup(self, _setup, cfs_cat):
        self.app = cfs_cat
        self.url = reverse("cat:cfs-schedule-add", kwargs={"cat_pk": self.app.pk})
        self.app.cfs_template.schedules.all().delete()

    def test_post_only(self):
        response = self.exporter_client.get(self.url)
        assert response.status_code == HTTPStatus.METHOD_NOT_ALLOWED

    def test_permission(self):
        response = self.importer_client.post(self.url)
        assert response.status_code == HTTPStatus.FORBIDDEN

        response = self.exporter_two_client.post(self.url)
        assert response.status_code == HTTPStatus.FORBIDDEN

    def test_post(self):
        assert self.app.cfs_template.schedules.count() == 0
        response = self.exporter_client.post(self.url)
        assert response.status_code == HTTPStatus.FOUND

        self.app.refresh_from_db()
        assert self.app.cfs_template.schedules.count() == 1


class TestCFSScheduleTemplateCopyView(AuthTestCase):
    @pytest.fixture(autouse=True)
    def setup(self, _setup, cfs_cat):
        self.app = cfs_cat
        self.url = reverse(
            "cat:cfs-schedule-copy",
            kwargs={
                "cat_pk": self.app.pk,
                "schedule_template_pk": self.app.cfs_template.schedules.first().pk,
            },
        )
        schedule = self.app.cfs_template.schedules.first()
        schedule.exporter_status = "MANUFACTURER"
        schedule.brand_name_holder = YesNoChoices.yes
        schedule.save()

        schedule.products.create(product_name="product 1")
        schedule.products.create(product_name="product 2")
        schedule.products.create(product_name="product 3")

    def test_post_only(self):
        response = self.exporter_client.get(self.url)
        assert response.status_code == HTTPStatus.METHOD_NOT_ALLOWED

    def test_permission(self):
        response = self.importer_client.post(self.url)
        assert response.status_code == HTTPStatus.FORBIDDEN

        response = self.exporter_two_client.post(self.url)
        assert response.status_code == HTTPStatus.FORBIDDEN

    def test_post(self):
        assert self.app.cfs_template.schedules.count() == 1
        initial_schedule = self.app.cfs_template.schedules.first()
        assert initial_schedule.products.all().count() == 3

        response = self.exporter_client.post(self.url)
        assert response.status_code == HTTPStatus.FOUND

        self.app.refresh_from_db()
        assert self.app.cfs_template.schedules.count() == 2

        copied_schedule = self.app.cfs_template.schedules.last()
        assert initial_schedule.pk != copied_schedule.pk
        assert copied_schedule.products.all().count() == 3

        assert copied_schedule.exporter_status == "MANUFACTURER"
        assert copied_schedule.brand_name_holder == YesNoChoices.yes


class TestCFSScheduleTemplateDeleteView(AuthTestCase):
    @pytest.fixture(autouse=True)
    def setup(self, _setup, cfs_cat):
        self.app = cfs_cat
        self.url = reverse(
            "cat:cfs-schedule-delete",
            kwargs={
                "cat_pk": self.app.pk,
                "schedule_template_pk": self.app.cfs_template.schedules.first().pk,
            },
        )

    def test_post_only(self):
        response = self.exporter_client.get(self.url)
        assert response.status_code == HTTPStatus.METHOD_NOT_ALLOWED

    def test_permission(self):
        response = self.importer_client.post(self.url)
        assert response.status_code == HTTPStatus.FORBIDDEN

        response = self.exporter_two_client.post(self.url)
        assert response.status_code == HTTPStatus.FORBIDDEN

    def test_post(self):
        assert self.app.cfs_template.schedules.count() == 1
        response = self.exporter_client.post(self.url)
        assert response.status_code == HTTPStatus.FOUND

        self.app.refresh_from_db()
        assert self.app.cfs_template.schedules.count() == 0


class TestCFSScheduleTemplateProductCreateView(AuthTestCase):
    @pytest.fixture(autouse=True)
    def setup(self, _setup, cfs_cat):
        self.app = cfs_cat
        self.url = reverse(
            "cat:cfs-schedule-product-create",
            kwargs={
                "cat_pk": self.app.pk,
                "schedule_template_pk": cfs_cat.cfs_template.schedules.first().pk,
            },
        )

    def test_permission(self):
        response = self.importer_client.get(self.url)
        assert response.status_code == HTTPStatus.FORBIDDEN

        response = self.exporter_two_client.get(self.url)
        assert response.status_code == HTTPStatus.FORBIDDEN

        response = self.exporter_client.get(self.url)
        assert response.status_code == HTTPStatus.OK

    def test_post(self):
        assert self.app.cfs_template.schedules.first().products.count() == 0

        form_data = {"product_name": "Test product 1"}
        response = self.exporter_client.post(self.url, data=form_data)
        assert response.status_code == HTTPStatus.FOUND

        schedule = self.app.cfs_template.schedules.first()
        assert schedule.products.count() == 1

        product = schedule.products.first()
        assert product.product_name == "Test product 1"


class TestCFSScheduleTemplateProductCreateMultipleView(AuthTestCase):
    @pytest.fixture(autouse=True)
    def setup(self, _setup, cfs_cat):
        self.app = cfs_cat
        self.schedule = cfs_cat.cfs_template.schedules.first()

        self.url = reverse(
            "cat:cfs-schedule-product-create-multiple",
            kwargs={
                "cat_pk": self.app.pk,
                "schedule_template_pk": self.schedule.pk,
            },
        )

    def test_permission(self):
        response = self.importer_client.get(self.url)
        assert response.status_code == HTTPStatus.FORBIDDEN

        response = self.exporter_two_client.get(self.url)
        assert response.status_code == HTTPStatus.FORBIDDEN

        response = self.exporter_client.get(self.url)
        assert response.status_code == HTTPStatus.OK

    def test_get(self):
        response = self.exporter_client.get(self.url)
        assert response.status_code == HTTPStatus.OK

        context = response.context
        assert context["formset"] is not None
        assert context["page_title"] == "Add Products"

    def test_post(self):
        assert self.schedule.products.count() == 0

        form_data = {
            "products-TOTAL_FORMS": "5",
            "products-INITIAL_FORMS": "0",
            "products-MIN_NUM_FORMS": "0",
            "products-MAX_NUM_FORMS": "1000",
            "products-0-id": "",
            "products-0-product_name": "Test product 1",
            "products-1-id": "",
            "products-1-product_name": "Test product 2",
            "products-2-id": "",
            "products-2-product_name": "Test product 3",
            "products-3-id": "",
            "products-3-product_name": "Test product 4",
            "products-4-id": "",
            "products-4-product_name": "Test product 5",
        }

        response = self.exporter_client.post(self.url, data=form_data)
        assert response.status_code == HTTPStatus.FOUND

        self.schedule.refresh_from_db()
        assert self.schedule.products.count() == 5
        assert list(self.schedule.products.all().values_list("product_name", flat=True)) == [
            "Test product 1",
            "Test product 2",
            "Test product 3",
            "Test product 4",
            "Test product 5",
        ]


class TestCFSScheduleTemplateProductUpdateView(AuthTestCase):
    @pytest.fixture(autouse=True)
    def setup(self, _setup, cfs_cat):
        self.app = cfs_cat
        self.schedule = self.app.cfs_template.schedules.first()
        self.product = self.schedule.products.create(product_name="Test product 1")

        self.url = reverse(
            "cat:cfs-schedule-product-update",
            kwargs={
                "cat_pk": self.app.pk,
                "schedule_template_pk": self.schedule.pk,
                "product_template_pk": self.product.pk,
            },
        )

    def test_permission(self):
        response = self.importer_client.get(self.url)
        assert response.status_code == HTTPStatus.FORBIDDEN

        response = self.exporter_two_client.get(self.url)
        assert response.status_code == HTTPStatus.FORBIDDEN

        response = self.exporter_client.get(self.url)
        assert response.status_code == HTTPStatus.OK

    def test_post(self):
        form_data = {
            "product_name": "Test product 1 updated",
        }
        response = self.exporter_client.post(self.url, data=form_data)
        assert response.status_code == HTTPStatus.FOUND

        self.product.refresh_from_db()
        assert self.product.product_name == "Test product 1 updated"

    def test_post_biocidal(self):
        self.schedule.legislations.add(ProductLegislation.objects.filter(is_biocidal=True).first())

        form_data = {
            # Product form
            "product_name": "Test product 1 updated",
            # Product types formset
            "pt_-TOTAL_FORMS": "3",
            "pt_-INITIAL_FORMS": "0",
            "pt_-MIN_NUM_FORMS": "0",
            "pt_-MAX_NUM_FORMS": "1000",
            "pt_-0-id": "",
            "pt_-0-product_type_number": "1",
            "pt_-1-id": "",
            "pt_-1-product_type_number": "2",
            "pt_-2-id": "",
            "pt_-2-product_type_number": "3",
            # Active ingredients formset
            "ai_-TOTAL_FORMS": "3",
            "ai_-INITIAL_FORMS": "0",
            "ai_-MIN_NUM_FORMS": "0",
            "ai_-MAX_NUM_FORMS": "1000",
            "ai_-0-id": "",
            "ai_-0-name": "Name 1",
            "ai_-0-cas_number": "111-11-1111",
            "ai_-1-id": "",
            "ai_-1-name": "Name 2",
            "ai_-1-cas_number": "222-22-2222",
            "ai_-2-id": "",
            "ai_-2-name": "Name 3",
            "ai_-2-cas_number": "333-33-3333",
        }

        response = self.exporter_client.post(self.url, data=form_data)
        assert response.status_code == HTTPStatus.FOUND

        self.product.refresh_from_db()
        assert self.product.product_name == "Test product 1 updated"

        assert list(
            self.product.product_type_numbers.all().values_list("product_type_number", flat=True)
        ) == [1, 2, 3]

        assert list(self.product.active_ingredients.all().values("name", "cas_number")) == [
            {"name": "Name 1", "cas_number": "111-11-1111"},
            {"name": "Name 2", "cas_number": "222-22-2222"},
            {"name": "Name 3", "cas_number": "333-33-3333"},
        ]


class TestCFSScheduleTemplateProductDeleteView(AuthTestCase):
    @pytest.fixture(autouse=True)
    def setup(self, _setup, cfs_cat):
        self.app = cfs_cat
        self.schedule = self.app.cfs_template.schedules.first()
        self.product = self.schedule.products.create(product_name="Test product 1")

        self.url = reverse(
            "cat:cfs-schedule-product-delete",
            kwargs={
                "cat_pk": self.app.pk,
                "schedule_template_pk": self.schedule.pk,
                "product_template_pk": self.product.pk,
            },
        )

    def test_post_only(self):
        response = self.exporter_client.get(self.url)
        assert response.status_code == HTTPStatus.METHOD_NOT_ALLOWED

    def test_permission(self):
        response = self.importer_client.post(self.url)
        assert response.status_code == HTTPStatus.FORBIDDEN

        response = self.exporter_two_client.post(self.url)
        assert response.status_code == HTTPStatus.FORBIDDEN

    def test_post(self):
        assert self.schedule.products.count() == 1

        response = self.exporter_client.post(self.url)
        assert response.status_code == HTTPStatus.FOUND

        self.schedule.refresh_from_db()
        assert self.schedule.products.count() == 0


class TestCFSScheduleTemplateProductDownloadSpreadsheetView(AuthTestCase):
    @pytest.fixture(autouse=True)
    def setup(self, _setup, cfs_cat):
        self.app = cfs_cat
        self.schedule = self.app.cfs_template.schedules.first()

        self.url = reverse(
            "cat:cfs-schedule-product-download-template",
            kwargs={
                "cat_pk": self.app.pk,
                "schedule_template_pk": self.schedule.pk,
            },
        )

    def test_permission(self):
        response = self.importer_client.post(self.url)
        assert response.status_code == HTTPStatus.FORBIDDEN

        response = self.exporter_two_client.post(self.url)
        assert response.status_code == HTTPStatus.FORBIDDEN

    def test_get(self):
        sheet = self.get_products_spreadsheet(False)

        assert sheet.max_column == 1
        assert sheet.max_row == 1

        spreadsheet_rows = sheet.values
        header = next(spreadsheet_rows)

        assert list(header) == ["Product Name"]

    def test_get_biocide(self):
        self.schedule.legislations.add(ProductLegislation.objects.filter(is_biocidal=True).first())
        sheet = self.get_products_spreadsheet(True)

        assert sheet.max_column == 4
        assert sheet.max_row == 1

        spreadsheet_rows = sheet.values
        header = next(spreadsheet_rows)

        assert list(header) == [
            "Product Name",
            "Product Type Numbers (CSV)",
            "Active Ingredient Name",
            "Active Ingredient CAS Number",
        ]

    def get_products_spreadsheet(self, is_biocidal):
        response = self.exporter_client.get(self.url)
        assert response.status_code == HTTPStatus.OK

        assert response["Content-Type"] == (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert response["Content-Disposition"] == (
            f"attachment; filename=CFS Product Upload {'Biocide ' if is_biocidal else ''}Template.xlsx"
        )
        workbook = load_workbook(filename=io.BytesIO(response.content), read_only=True)
        assert workbook.sheetnames == ["CFS Products"]
        sheet = workbook["CFS Products"]

        return sheet


class TestCFSScheduleTemplateProductUploadSpreadsheetView(AuthTestCase):
    @pytest.fixture(autouse=True)
    def setup(self, _setup, cfs_cat):
        self.app = cfs_cat
        self.schedule = self.app.cfs_template.schedules.first()

        self.url = reverse(
            "cat:cfs-schedule-product-spreadsheet-upload",
            kwargs={
                "cat_pk": self.app.pk,
                "schedule_template_pk": self.schedule.pk,
            },
        )

    def test_post_only(self):
        response = self.exporter_client.get(self.url)
        assert response.status_code == HTTPStatus.METHOD_NOT_ALLOWED

    def test_permission(self):
        response = self.importer_client.post(self.url)
        assert response.status_code == HTTPStatus.FORBIDDEN

        response = self.exporter_two_client.post(self.url)
        assert response.status_code == HTTPStatus.FORBIDDEN

    @pytest.mark.parametrize("is_biocidal", [False, True])
    def test_post(self, is_biocidal):
        if is_biocidal:
            self.schedule.legislations.add(
                ProductLegislation.objects.filter(is_biocidal=True).first()
            )

        config = create_dummy_config(is_biocidal=is_biocidal)
        file_name = f"CFS Product Upload {'Biocide ' if is_biocidal else ''}Template.xlsx"
        file = create_dummy_xlsx_file(config, file_name)

        with mock.patch("web.domains.cat.views.delete_file_from_s3") as mock_s3:
            response = self.exporter_client.post(self.url, data={"file": file})

            mock_s3.assert_called_once_with(file_name)

        assert response.status_code == HTTPStatus.FOUND
        self.schedule.refresh_from_db()
        assert self.schedule.products.count() == 3
        assert self.schedule.products.filter(product_name="Product 1").count() == 1

        if is_biocidal:
            product_1 = self.schedule.products.get(product_name="Product 1")
            assert product_1.active_ingredients.count() == 2
            assert product_1.product_type_numbers.count() == 3
