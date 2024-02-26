import datetime as dt
import io
from http import HTTPStatus

import pytest
from django.core import mail
from django.test.client import Client
from django.utils import timezone
from guardian.shortcuts import remove_perm
from openpyxl import load_workbook
from pytest_django.asserts import assertRedirects, assertTemplateUsed

from web.domains.case.services import case_progress, document_pack
from web.domains.case.shared import ImpExpStatus
from web.flow import errors
from web.mail.constants import EmailTypes
from web.mail.url_helpers import get_case_view_url, get_validate_digital_signatures_url
from web.models import (
    CertificateOfManufactureApplication,
    ImportApplicationLicence,
    SILApplication,
    Task,
    WoodQuotaApplication,
)
from web.permissions import Perms
from web.sites import get_exporter_site_domain, get_importer_site_domain
from web.tests.helpers import CaseURLS, SearchURLS, check_gov_notify_email_was_sent
from web.utils.search.types import ExportResultRow, ImportResultRow


class TestSearchCasesView:
    @pytest.fixture(autouse=True)
    def _setup(self, importer_client, exporter_client, ilb_admin_client):
        self.import_url = SearchURLS.search_cases("import")
        self.export_url = SearchURLS.search_cases("export")

        self.importer_user_client = importer_client
        self.exporter_user_client = exporter_client
        self.ilb_admin_user_client = ilb_admin_client

    def test_permission(self):
        response = self.importer_user_client.get(self.import_url)
        assert response.status_code == HTTPStatus.OK

        response = self.exporter_user_client.get(self.import_url)
        assert response.status_code == HTTPStatus.FORBIDDEN

        response = self.exporter_user_client.get(self.export_url)
        assert response.status_code == HTTPStatus.OK

        response = self.importer_user_client.get(self.export_url)
        assert response.status_code == HTTPStatus.FORBIDDEN

        for search_url in [self.import_url, self.export_url]:
            response = self.ilb_admin_user_client.get(search_url)
            assert response.status_code == HTTPStatus.OK

    def test_can_search_for_a_case_import(self, completed_sil_app):
        form_data = {"case_ref": completed_sil_app.reference}

        results_url = SearchURLS.search_cases_get_results("import")
        response = self.importer_user_client.get(results_url, form_data)

        assert response.status_code == HTTPStatus.OK
        assert response.context["show_search_results"] is True
        assert response.context["total_rows"] == 1

        result: ImportResultRow = response.context["search_records"][0]
        assert result.case_status.case_reference == completed_sil_app.reference

    def test_can_search_for_a_case_export(self, completed_cfs_app):
        form_data = {"case_ref": completed_cfs_app.reference}

        results_url = SearchURLS.search_cases_get_results("export")
        response = self.exporter_user_client.get(results_url, form_data)

        assert response.status_code == HTTPStatus.OK
        assert response.context["show_search_results"] is True
        assert response.context["total_rows"] == 1

        result: ExportResultRow = response.context["search_records"][0]
        assert result.case_reference == completed_cfs_app.reference


class TestDownloadSpreadsheetView:
    @pytest.fixture(autouse=True)
    def _setup(self, importer_client, exporter_client, ilb_admin_client):
        self.import_download_url = SearchURLS.download_spreadsheet("import")
        self.export_download_url = SearchURLS.download_spreadsheet("export")

        self.importer_user_client = importer_client
        self.exporter_user_client = exporter_client
        self.ilb_admin_user_client = ilb_admin_client

    def test_permission(self):
        response = self.importer_user_client.post(self.import_download_url, data={})
        assert response.status_code == HTTPStatus.OK

        response = self.exporter_user_client.post(self.import_download_url, data={})
        assert response.status_code == HTTPStatus.FORBIDDEN

        response = self.exporter_user_client.post(self.export_download_url, data={})
        assert response.status_code == HTTPStatus.OK

        response = self.importer_user_client.post(self.export_download_url, data={})
        assert response.status_code == HTTPStatus.FORBIDDEN

        for search_url in [self.import_download_url, self.export_download_url]:
            response = self.ilb_admin_user_client.post(search_url, data={})
            assert response.status_code == HTTPStatus.OK

    def test_can_download_spreadsheet_import(self, completed_sil_app):
        form_data = {"case_ref": completed_sil_app.reference}
        response = self.importer_user_client.post(self.import_download_url, form_data)

        assert response.status_code == HTTPStatus.OK
        workbook = load_workbook(filename=io.BytesIO(response.content), read_only=True)

        assert workbook.sheetnames == ["Sheet 1"]

        sheet = workbook["Sheet 1"]

        assert sheet.max_column == 19
        assert sheet.max_row == 2

        spreadsheet_rows = sheet.values
        header = next(spreadsheet_rows)

        assert list(header) == [
            "Case Reference",
            "Applicant's Reference",
            "Licence Reference",
            "Licence Type",
            "Licence Start Date",
            "Licence End Date",
            "Application Type",
            "Application Sub-Type",
            "Case Status",
            "Chief Usage Status",
            "Submitted Date",
            "Importer",
            "Agent",
            "Application Contact",
            "Country of Origin",
            "Country of Consignment",
            "Shipping Year",
            "Goods Category",
            "Commodity Code(s)",
        ]

        applicant_refs = []

        # Iterate over the remaining data rows
        for row_data in spreadsheet_rows:
            applicant_refs.append(row_data[0])

        assert applicant_refs == [completed_sil_app.reference]

    def test_can_download_spreadsheet_export(self, completed_cfs_app):
        form_data = {"case_ref": completed_cfs_app.reference}
        response = self.exporter_user_client.post(self.export_download_url, form_data)

        assert response.status_code == HTTPStatus.OK
        workbook = load_workbook(filename=io.BytesIO(response.content), read_only=True)

        assert workbook.sheetnames == ["Sheet 1"]

        sheet = workbook["Sheet 1"]
        assert sheet.max_column == 10
        assert sheet.max_row == 2

        spreadsheet_rows = sheet.values
        header = next(spreadsheet_rows)

        assert list(header) == [
            "Case Reference",
            "Certificates",
            "Application Type",
            "Status",
            "Submitted Date",
            "Certificate Countries",
            "Countries of Manufacture",
            "Exporter",
            "Agent",
            "Application Contact",
        ]

        case_refs = []

        # Iterate over the remaining data rows
        for row_data in spreadsheet_rows:
            case_refs.append(row_data[0])

        assert case_refs == [completed_cfs_app.reference]


class TestReassignCaseOwnerView:
    def test_permission(self, importer_client, exporter_client, ilb_admin_client):
        importer_user_client = importer_client
        exporter_user_client = exporter_client

        url = SearchURLS.reassign_case_owner()

        response = importer_user_client.post(url, data={})
        assert response.status_code == HTTPStatus.FORBIDDEN

        response = exporter_user_client.post(url, data={})
        assert response.status_code == HTTPStatus.FORBIDDEN

        # 400 response is valid when form data is incorrect.
        # It shows the ILB user has permission.
        response = ilb_admin_client.post(url, data={})
        assert response.status_code == HTTPStatus.BAD_REQUEST

    def test_reassign_case_owner_import(
        self, fa_dfl_app_submitted, ilb_admin_client, ilb_admin_user, ilb_admin_two
    ):
        app = fa_dfl_app_submitted

        assert app.case_owner is None

        ilb_admin_client.post(CaseURLS.take_ownership(app.pk, "import"))
        app.refresh_from_db()
        assert app.case_owner == ilb_admin_user

        # Test reassignment
        form_data = {"assign_to": ilb_admin_two.pk, "applications": fa_dfl_app_submitted.pk}
        url = SearchURLS.reassign_case_owner("import")
        response = ilb_admin_client.post(url, form_data)

        assert response.status_code == HTTPStatus.NO_CONTENT

        app.refresh_from_db()

        assert app.case_owner == ilb_admin_two

    def test_reassign_case_owner_export(
        self, com_app_submitted, ilb_admin_client, ilb_admin_user, ilb_admin_two
    ):
        app = com_app_submitted

        assert app.case_owner is None

        ilb_admin_client.post(CaseURLS.take_ownership(app.pk, "export"))
        app.refresh_from_db()
        assert app.case_owner == ilb_admin_user

        # Test reassignment
        form_data = {"assign_to": ilb_admin_two.pk, "applications": com_app_submitted.pk}
        url = SearchURLS.reassign_case_owner("export")
        response = ilb_admin_client.post(url, form_data)

        assert response.status_code == HTTPStatus.NO_CONTENT

        app.refresh_from_db()

        assert app.case_owner == ilb_admin_two


class TestReopenApplicationViewImportApplication:
    client: Client
    wood_app: WoodQuotaApplication

    @pytest.fixture(autouse=True)
    def set_client(self, ilb_admin_client):
        self.client = ilb_admin_client

    @pytest.fixture(autouse=True)
    def set_app(self, wood_app_submitted, ilb_admin_user):
        self.wood_app = wood_app_submitted

        # End the PROCESS task as we are testing reopening the application
        task = case_progress.get_expected_task(self.wood_app, Task.TaskType.PROCESS)
        task.is_active = False
        task.finished = timezone.now()
        task.owner = ilb_admin_user
        task.save()

    def test_permission(self, importer_client, exporter_client):
        url = SearchURLS.reopen_case(self.wood_app.pk)

        response = importer_client.get(url)
        assert response.status_code == HTTPStatus.FORBIDDEN

        response = exporter_client.get(url)
        assert response.status_code == HTTPStatus.FORBIDDEN

        # self.client is an ILB admin user
        self.wood_app.status = ImpExpStatus.STOPPED
        self.wood_app.save()
        resp = self.client.post(url)

        self._check_valid_response(resp, self.wood_app)

    def test_reopen_application_when_stopped(self):
        self.wood_app.status = ImpExpStatus.STOPPED
        self.wood_app.save()

        url = SearchURLS.reopen_case(application_pk=self.wood_app.pk)
        resp = self.client.post(url)

        self._check_valid_response(resp, self.wood_app)

    def test_reopen_application_when_withdrawn(self):
        self.wood_app.status = ImpExpStatus.WITHDRAWN
        self.wood_app.save()

        url = SearchURLS.reopen_case(application_pk=self.wood_app.pk)
        resp = self.client.post(url)

        self._check_valid_response(resp, self.wood_app)

    def test_reopen_application_when_processing_errors(self):
        with pytest.raises(expected_exception=errors.ProcessStateError):
            url = SearchURLS.reopen_case(application_pk=self.wood_app.pk)
            self.client.post(url)
            self._check_email_is_not_sent()

    def test_reopen_application_unsets_caseworker(self, ilb_admin_user):
        self.wood_app.status = ImpExpStatus.STOPPED
        self.wood_app.case_owner = ilb_admin_user
        self.wood_app.save()

        url = SearchURLS.reopen_case(application_pk=self.wood_app.pk)
        resp = self.client.post(url)

        self._check_valid_response(resp, self.wood_app)
        assert self.wood_app.case_owner is None

    def _check_valid_response(self, resp, application):
        assert resp.status_code == HTTPStatus.NO_CONTENT
        application.refresh_from_db()

        case_progress.check_expected_status(application, [application.Statuses.SUBMITTED])
        case_progress.check_expected_task(application, Task.TaskType.PROCESS)

        self._check_email_is_sent(application)

    def _check_email_is_sent(self, application):
        check_gov_notify_email_was_sent(
            1,
            ["I1_main_contact@example.com"],  # /PS-IGNORE
            EmailTypes.APPLICATION_REOPENED,
            {
                "reference": application.reference,
                "validate_digital_signatures_url": get_validate_digital_signatures_url(
                    full_url=True
                ),
                "application_url": get_case_view_url(application, get_importer_site_domain()),
                "icms_url": get_importer_site_domain(),
            },
        )

    def _check_email_is_not_sent(self):
        outbox = mail.outbox
        assert len(outbox) == 0


class TestReopenApplicationViewExportApplication:
    client: Client
    app: CertificateOfManufactureApplication

    @pytest.fixture(autouse=True)
    def set_client(self, ilb_admin_client):
        self.client = ilb_admin_client

    @pytest.fixture(autouse=True)
    def set_app(self, com_app_submitted, ilb_admin_user):
        self.app = com_app_submitted

        # End the PROCESS task as we are testing reopening the application
        task = case_progress.get_expected_task(self.app, Task.TaskType.PROCESS)
        task.is_active = False
        task.finished = timezone.now()
        task.owner = ilb_admin_user
        task.save()

    def test_permission(self, importer_client, exporter_client):
        url = SearchURLS.reopen_case(self.app.pk, case_type="export")

        response = importer_client.get(url)
        assert response.status_code == HTTPStatus.FORBIDDEN

        response = exporter_client.get(url)
        assert response.status_code == HTTPStatus.FORBIDDEN

        # self.client is an ILB admin user
        self.app.status = ImpExpStatus.STOPPED
        self.app.save()
        resp = self.client.post(url)

        self._check_valid_response(resp, self.app)

    def test_reopen_application_when_stopped(self):
        self.app.status = ImpExpStatus.STOPPED
        self.app.save()

        url = SearchURLS.reopen_case(application_pk=self.app.pk)
        resp = self.client.post(url)

        self._check_valid_response(resp, self.app)

    def test_reopen_application_when_withdrawn(self):
        self.app.status = ImpExpStatus.WITHDRAWN
        self.app.save()

        url = SearchURLS.reopen_case(application_pk=self.app.pk)
        resp = self.client.post(url)

        self._check_valid_response(resp, self.app)

    def test_reopen_application_when_processing_errors(self):
        with pytest.raises(expected_exception=errors.ProcessStateError):
            url = SearchURLS.reopen_case(application_pk=self.app.pk)
            self.client.post(url)
            self._check_email_is_not_sent()

    def test_reopen_application_unsets_caseworker(self, ilb_admin_user):
        self.app.status = ImpExpStatus.STOPPED
        self.app.case_owner = ilb_admin_user
        self.app.save()

        url = SearchURLS.reopen_case(application_pk=self.app.pk)
        resp = self.client.post(url)

        self._check_valid_response(resp, self.app)
        assert self.app.case_owner is None

    def _check_valid_response(self, resp, application):
        assert resp.status_code == HTTPStatus.NO_CONTENT
        application.refresh_from_db()

        case_progress.check_expected_status(application, [application.Statuses.SUBMITTED])
        case_progress.check_expected_task(application, Task.TaskType.PROCESS)

        self._check_email_is_sent(application)

    def _check_email_is_sent(self, application):
        check_gov_notify_email_was_sent(
            1,
            ["E1_main_contact@example.com"],  # /PS-IGNORE
            EmailTypes.APPLICATION_REOPENED,
            {
                "reference": application.reference,
                "validate_digital_signatures_url": get_validate_digital_signatures_url(
                    full_url=True
                ),
                "application_url": get_case_view_url(application, get_exporter_site_domain()),
                "icms_url": get_exporter_site_domain(),
            },
        )

    def _check_email_is_not_sent(self):
        outbox = mail.outbox
        assert len(outbox) == 0


class TestRequestVariationUpdateView:
    client: Client
    wood_app: WoodQuotaApplication

    @pytest.fixture(autouse=True)
    def set_client(self, ilb_admin_client):
        self.client = ilb_admin_client

    @pytest.fixture(autouse=True)
    def set_app(self, wood_app_submitted, ilb_admin_user, importer_one_contact):
        self.wood_app = wood_app_submitted
        self.wood_app.status = ImpExpStatus.COMPLETED
        self.wood_app.save()

        # A completed app must have an active licence
        draft_pack = document_pack.pack_draft_get(self.wood_app)
        draft_pack.issue_paper_licence_only = True
        draft_pack.licence_start_date = dt.date(2020, 6, 14)
        draft_pack.licence_end_date = dt.date(2023, 9, 15)
        draft_pack.save()

        document_pack.pack_draft_set_active(self.wood_app)

        # End the PROCESS task as we are testing with a completed application
        task = case_progress.get_expected_task(self.wood_app, Task.TaskType.PROCESS)
        task.is_active = False
        task.finished = timezone.now()
        task.owner = ilb_admin_user
        task.save()

        self.wood_app.cleared_by.add(importer_one_contact)

    def test_permission(
        self, importer_client, exporter_client, ilb_admin_client, importer_one_contact
    ):
        ilb_admin_user_client = ilb_admin_client

        url = SearchURLS.request_variation(self.wood_app.pk)

        response = importer_client.get(url)
        assert response.status_code == HTTPStatus.OK

        # Removing the edit object permission should prevent the importer contact from
        # attempting to perform a variation request.
        remove_perm(Perms.obj.importer.edit, importer_one_contact, self.wood_app.importer)
        response = importer_client.get(url)
        assert response.status_code == HTTPStatus.FORBIDDEN

        response = exporter_client.get(url)
        assert response.status_code == HTTPStatus.FORBIDDEN

        response = ilb_admin_user_client.get(url)
        assert response.status_code == HTTPStatus.OK

        remove_perm(Perms.obj.importer.edit, importer_one_contact, self.wood_app.importer)

    def test_get_search_url(self):
        url = SearchURLS.request_variation(self.wood_app.pk)

        # Test referrer query params are remembered
        referrer = "http://caseworker:8080/case/import/search/standard/results/?status=COMPLETED&decision=APPROVE"
        resp = self.client.get(url, HTTP_REFERER=referrer)

        expected = "/case/import/search/standard/results/?status=COMPLETED&decision=APPROVE"
        assert resp.context["search_results_url"] == expected

        # Test referrer set without query params
        referrer = "http://caseworker:8080/case/import/search/standard/results/"
        resp = self.client.get(url, HTTP_REFERER=referrer)

        expected = "/case/import/search/standard/results/"
        assert resp.context["search_results_url"] == expected

    def test_get_search_url_no_referrer(self):
        url = SearchURLS.request_variation(self.wood_app.pk)
        resp = self.client.get(url)

        expected = "/case/import/search/standard/results/?case_status=VARIATION_REQUESTED"
        assert resp.context["search_results_url"] == expected

    def test_post_updates_status(self, ilb_admin_user, importer_one_contact):
        url = SearchURLS.request_variation(self.wood_app.pk)

        live_licence = document_pack.pack_active_get(self.wood_app)
        assert live_licence.status == ImportApplicationLicence.Status.ACTIVE

        form_data = {
            "what_varied": "What was varied",
            "why_varied": "Why was it varied",
            "when_varied": "01-Jan-2021",
        }

        # Set this fields as reopening a case should clear them.
        self.wood_app.case_owner = ilb_admin_user
        self.wood_app.variation_decision = WoodQuotaApplication.REFUSE
        self.wood_app.variation_refuse_reason = "test value"

        assert self.wood_app.cleared_by.filter(pk=importer_one_contact.pk).exists()

        resp = self.client.post(url, data=form_data)
        default_redirect_url = (
            "/case/import/search/standard/results/?case_status=VARIATION_REQUESTED"
        )

        assertRedirects(resp, default_redirect_url, 302)
        self.wood_app.refresh_from_db()

        assert not self.wood_app.case_owner
        assert not self.wood_app.variation_decision
        assert not self.wood_app.variation_refuse_reason

        case_progress.check_expected_status(self.wood_app, [ImpExpStatus.VARIATION_REQUESTED])
        case_progress.check_expected_task(self.wood_app, Task.TaskType.PROCESS)

        # Check the application's latest licence status is draft
        latest_licence = document_pack.pack_draft_get(self.wood_app)
        assert latest_licence.status == ImportApplicationLicence.Status.DRAFT

        # All the old values should have been copied over
        assert live_licence.licence_start_date == latest_licence.licence_start_date
        assert live_licence.licence_end_date == latest_licence.licence_end_date
        assert live_licence.issue_paper_licence_only == latest_licence.issue_paper_licence_only

        # Check cleared by has been cleared
        assert self.wood_app.cleared_by.count() == 0


class TestRequestVariationOpenRequestView:
    client: Client
    app: "CertificateOfManufactureApplication"

    @pytest.fixture(autouse=True)
    def set_client(self, ilb_admin_client):
        self.client = ilb_admin_client

    @pytest.fixture(autouse=True)
    def set_app(self, com_app_submitted, ilb_admin_user, exporter_one_contact):
        self.app = com_app_submitted
        self.app.status = ImpExpStatus.COMPLETED
        # A completed app would have a case owner (to test it gets cleared)
        self.app.case_owner = ilb_admin_user
        self.app.save()

        # End the PROCESS task as we are testing with a completed application
        task = case_progress.get_expected_task(self.app, Task.TaskType.PROCESS)
        task.is_active = False
        task.finished = timezone.now()
        task.owner = ilb_admin_user
        task.save()

        self.app.cleared_by.add(exporter_one_contact)

    def test_permission(
        self,
        importer_client,
        exporter_client,
        ilb_admin_client,
        wood_app_submitted,
    ):
        url = SearchURLS.open_variation(self.app.pk)

        response = importer_client.get(url)
        assert response.status_code == HTTPStatus.FORBIDDEN

        response = exporter_client.get(url)
        assert response.status_code == HTTPStatus.FORBIDDEN

        # self.client is an ILB admin user
        response = self.client.get(url)
        assert response.status_code == HTTPStatus.OK

    def test_post_updates_status(self, ilb_admin_user, exporter_one_contact):
        url = SearchURLS.open_variation(self.app.pk)

        form_data = {"what_varied": "What was varied"}

        assert self.app.cleared_by.filter(pk=exporter_one_contact.pk).exists()

        response = self.client.post(url, data=form_data)
        default_redirect_url = (
            "/case/export/search/standard/results/?case_status=VARIATION_REQUESTED"
        )

        assertRedirects(response, default_redirect_url, 302)
        self.app.refresh_from_db()

        assert not self.app.case_owner

        case_progress.check_expected_status(self.app, [ImpExpStatus.VARIATION_REQUESTED])
        case_progress.check_expected_task(self.app, Task.TaskType.PROCESS)

        # Check cleared by has been cleared
        assert self.app.cleared_by.count() == 0


class TestRevokeCaseView:
    client: Client
    app: "SILApplication"
    url: str

    @pytest.fixture(autouse=True)
    def setup(self, ilb_admin_client, completed_sil_app):
        self.client = ilb_admin_client
        self.app = completed_sil_app
        self.url = SearchURLS.revoke_licence(self.app.pk)

    def test_permission(self, importer_client, exporter_client):
        response = self.client.get(self.url)
        assert response.status_code == HTTPStatus.OK

        response = importer_client.get(self.url)
        assert response.status_code == HTTPStatus.FORBIDDEN

        response = exporter_client.get(self.url)
        assert response.status_code == HTTPStatus.FORBIDDEN

    def test_revoke_certificate_with_send_email(self, completed_cfs_app):
        app = completed_cfs_app
        url = SearchURLS.revoke_licence(app.pk, case_type="export")
        response = self.client.get(url)
        assert response.status_code == HTTPStatus.OK

        assertTemplateUsed(response, "web/domains/case/manage/revoke-certificate.html")
        assert "Certificate Revocation" in response.content.decode()

        form_data = {"send_email": "on", "reason": "test reason"}
        resp = self.client.post(url, data=form_data, follow=True)
        assert resp.status_code == HTTPStatus.OK

        app.refresh_from_db()
        assert app.status == ImpExpStatus.REVOKED

        pack = document_pack.pack_revoked_get(app)
        assert pack.revoke_reason == "test reason"
        assert pack.revoke_email_sent is True

        year = timezone.now().year
        check_gov_notify_email_was_sent(
            1,
            ["E1_main_contact@example.com"],  # /PS-IGNORE,
            EmailTypes.CERTIFICATE_REVOKED,
            {
                "reference": app.reference,
                "validate_digital_signatures_url": get_validate_digital_signatures_url(
                    full_url=True
                ),
                "application_url": get_case_view_url(app, get_exporter_site_domain()),
                "icms_url": get_exporter_site_domain(),
                "certificate_references": f"CFS/{year}/00001,CFS/{year}/00002",
            },
        )

    def test_revoke_licence_with_send_email(self):
        # check what is in the context on the initial page load
        response = self.client.get(self.url)
        assert response.status_code == HTTPStatus.OK
        assertTemplateUsed(response, "web/domains/case/manage/revoke-licence.html")
        assert "Licence Revocation" in response.content.decode()

        assert response.context["active_pack"] == document_pack.pack_active_get(self.app)
        assert response.context["process"] == self.app
        assert response.context["search_results_url"].startswith(
            "/case/import/search/standard/results/?case_ref=IMA%2F"
        )

        form_data = {"send_email": "on", "reason": "test reason"}
        resp = self.client.post(self.url, data=form_data, follow=True)
        assert resp.status_code == HTTPStatus.OK

        self.app.refresh_from_db()
        assert self.app.status == ImpExpStatus.REVOKED

        pack = document_pack.pack_revoked_get(self.app)
        licence = document_pack.doc_ref_licence_get(pack)
        assert pack.revoke_reason == "test reason"
        assert pack.revoke_email_sent is True

        check_gov_notify_email_was_sent(
            1,
            ["I1_main_contact@example.com"],  # /PS-IGNORE,
            EmailTypes.LICENCE_REVOKED,
            {
                "reference": self.app.reference,
                "validate_digital_signatures_url": get_validate_digital_signatures_url(
                    full_url=True
                ),
                "application_url": get_case_view_url(self.app, get_importer_site_domain()),
                "icms_url": get_importer_site_domain(),
                "licence_number": licence.reference,
            },
        )

        # check what is in the context on the revoked licence page after revoking
        response = self.client.get(self.url)
        assert response.context["active_pack"] is None

        # Check the previous data is preserved and the form is readonly
        form = response.context["form"]
        assert form.initial == {"reason": "test reason", "send_email": True}
        for f in form.fields:
            assert form.fields[f].disabled

        assert response.context["process"] == self.app
        assert response.context["search_results_url"].startswith(
            "/case/import/search/standard/results/?case_ref=IMA%2F"
        )

    def test_revoke_licence_with_no_email(self):
        form_data = {"reason": "test reason"}
        resp = self.client.post(self.url, data=form_data, follow=True)
        assert resp.status_code == HTTPStatus.OK

        self.app.refresh_from_db()
        assert self.app.status == ImpExpStatus.REVOKED

        pack = document_pack.pack_revoked_get(self.app)
        assert pack.revoke_reason == "test reason"
        assert pack.revoke_email_sent is False
        assert len(mail.outbox) == 0
