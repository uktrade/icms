import io
from collections import OrderedDict
from dataclasses import dataclass, field

from django.core.files.base import File
from django.http import HttpResponse
from openpyxl import load_workbook

from web.domains.case.app_copy import (
    ApplicationCopy,
    CFSApplicationCopy,
    COMApplicationCopy,
    GMPApplicationCopy,
)
from web.domains.case.export.forms import (
    CFSActiveIngredientForm,
    CFSProductForm,
    CFSProductTypeForm,
)
from web.domains.cat.forms import (
    CFSActiveIngredientTemplateForm,
    CFSProductTemplateForm,
    CFSProductTypeTemplateForm,
)
from web.models import (
    CertificateApplicationTemplate,
    CertificateOfFreeSaleApplication,
    CertificateOfFreeSaleApplicationTemplate,
    CertificateOfGoodManufacturingPracticeApplication,
    CertificateOfGoodManufacturingPracticeApplicationTemplate,
    CertificateOfManufactureApplication,
    CertificateOfManufactureApplicationTemplate,
    CFSProduct,
    CFSProductTemplate,
    CFSSchedule,
    CFSScheduleTemplate,
    ExportApplication,
    ExportApplicationType,
    Exporter,
    Office,
    User,
)
from web.utils.spreadsheet import XlsxSheetConfig, generate_xlsx_file


class CustomError(Exception):
    pass


@dataclass
class ProductData:
    product_type_numbers: list[int] = field(default_factory=list)
    ingredient_names: list[str] = field(default_factory=list)
    cas_numbers: list[str] = field(default_factory=list)


def get_product_spreadsheet_response(schedule: CFSSchedule | CFSScheduleTemplate) -> HttpResponse:
    is_biocidal = schedule.is_biocidal()
    mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    response = HttpResponse(content_type=mime_type)
    workbook = generate_product_template_xlsx(is_biocidal)
    response.write(workbook)

    filename = "CFS Product Upload Template.xlsx"

    if is_biocidal:
        filename = "CFS Product Upload Biocide Template.xlsx"

    response["Content-Disposition"] = f"attachment; filename={filename}"

    return response


def generate_product_template_xlsx(is_biocidal: bool = False) -> bytes:
    """Generates the schedule products xslx template for download"""

    header_data = _get_header(is_biocidal)
    config = XlsxSheetConfig()
    config.header.data = header_data
    config.header.styles = {"bold": True}
    config.column_width = 25
    config.sheet_name = "CFS Products"

    xlsx_data = generate_xlsx_file([config])
    return xlsx_data


def process_products_file(products_file: File, schedule: CFSSchedule | CFSScheduleTemplate) -> int:
    """Processes the uploaded xlsx file and save the products to the schedule.

    Raises an exception on any validation failure.
    Returns the count of the products which were processed.
    """

    is_biocidal = schedule.is_biocidal()

    products_data = _extract_product_data(products_file, is_biocidal)
    product_count = len(products_data)

    for product_name, product_data in products_data.items():
        product = _process_product(schedule, product_name)

        _process_product_type_numbers(product, product_data)
        _process_ingredients(product, product_data)

    return product_count


def _get_header(is_biocidal: bool = False) -> list[str]:
    """Generates the header row for the xlsx file"""

    if is_biocidal:
        return [
            "Product Name",
            "Product Type Numbers (CSV)",
            "Active Ingredient Name",
            "Active Ingredient CAS Number",
        ]

    return ["Product Name"]


def _extract_sheet_header(sheet, is_biocidal: bool = False) -> list[str]:
    """Get the expected header for xlsx and compare it to header in the first row of data"""

    header = _get_header(is_biocidal)
    header_row = [col.strip() if col else "" for col in next(sheet.values)]

    # Check the number of columns with data in xlsx file matches the length of the header
    if "" in header_row:
        raise CustomError(
            "Number of columns with data do not match the number of columns in the header"
        )

    # Check the header in the xlsx file matches the expected template header
    if header != header_row:
        raise CustomError("Spreadsheet header does not match the template header")

    return header


def _extract_product_data(products_file: File, is_biocidal: bool = False) -> dict[str, ProductData]:
    """Iterates over the rows in the products xlsx and extracts the product data from each row.

    Combines product type numbers and ingredients for the each product.
    Returns the data as an OrderedDict.
    """

    if products_file.multiple_chunks():
        # If there is more than one chunk, the file is too big
        raise CustomError("File too large to process")

    data: OrderedDict = OrderedDict()
    chunk = next(products_file.chunks())
    workbook = load_workbook(filename=io.BytesIO(chunk), read_only=True, data_only=True)

    try:
        if "CFS Products" not in workbook.sheetnames:
            raise CustomError("Cannot find sheet with name 'CFS Products' in file")

        sheet = workbook["CFS Products"]
        header = _extract_sheet_header(sheet, is_biocidal)

        for row, row_values in enumerate(sheet.values, start=1):
            # Skip the header row
            if row == 1:
                continue

            # Clean values in row by stripping whitespace
            values = [
                value and str(value).strip() or ""
                for value in row_values  # type:ignore[union-attr]
            ]

            # Create a dict of values using column name
            row_data = dict(zip(header, values))
            row_data["row"] = str(row)

            product_name = row_data.get("Product Name")

            if not product_name:
                raise CustomError(f"Data missing in column 'Product Name' - line {row}")

            product = data.get(product_name, ProductData())

            # There will be only the product name column in the sheet for non biocidal legislation
            if not is_biocidal:
                data[product_name] = product

                continue

            _add_product_type_numbers_to_product_data(row_data, product)
            _add_ingredient_to_product_data(row_data, product)

            data[product_name] = product

        workbook.close()

    except Exception as err:
        # Catch the exception to explicitly close the workbook before raising
        workbook.close()
        raise err

    return data


def _add_product_type_numbers_to_product_data(
    row_data: dict[str, str], product: ProductData
) -> None:
    """Gets and validates product type numbers data from the row and adds to ProductData.

    Appends the product types numbers to ProductData.product_type_numbers.
    """

    data = row_data.get("Product Type Numbers (CSV)", "")
    product_name = row_data["Product Name"]
    row = row_data["row"]

    for val in data.split(","):
        try:
            # Integers are stored as floats in xlsx files
            product_type = int(float(val))

            if product_type not in product.product_type_numbers:
                product.product_type_numbers.append(product_type)

        except ValueError:
            raise CustomError(
                f"Product type number '{val}' for product '{product_name}' is not a number - line {row}"
            )


def _add_ingredient_to_product_data(
    row_data: dict[str, str],
    product: ProductData,
) -> None:
    """Gets and validates ingredient data from the row and adds to ProductData.

    Appends ingredient name to ProductData.ingredient_names.
    Appends cas number to ProductData.cas_numbers.
    """

    ingredient_name = row_data.get("Active Ingredient Name")
    cas_number = row_data.get("Active Ingredient CAS Number")
    product_name = row_data["Product Name"]
    row = row_data["row"]

    if not ingredient_name:
        raise CustomError(f"Ingredient name missing - line {row}")

    if not cas_number:
        raise CustomError(f"CAS number missing - line {row}")

    if ingredient_name in product.ingredient_names:
        raise CustomError(
            f"Ingredient name '{ingredient_name}' duplicated for product '{product_name}' - line {row}"
        )

    if cas_number in product.cas_numbers:
        raise CustomError(
            f"CAS number '{cas_number}' duplicated for product '{product_name}' - line {row}"
        )

    product.ingredient_names.append(ingredient_name)
    product.cas_numbers.append(cas_number)


def _process_product(schedule: CFSSchedule | CFSScheduleTemplate, product_name: str) -> CFSProduct:
    """Gets existing product or saves new product using CFSProductForm and returns product"""

    existing_product = schedule.products.filter(product_name__iexact=product_name).first()

    if existing_product:
        return existing_product

    if isinstance(schedule, CFSSchedule):
        product_form_cls = CFSProductForm
    else:
        product_form_cls = CFSProductTemplateForm

    form = product_form_cls(data={"product_name": product_name}, schedule=schedule)

    if not form.is_valid():
        errors = form.errors

        msg = f"Product '{product_name}' has error"

        if "product_name" in errors:
            msg += f" - {errors['product_name'][0]}"

        raise CustomError(msg)

    return form.save()


def _process_product_type_numbers(
    product: CFSProduct | CFSProductTemplate, product_data: ProductData
) -> None:
    """Gets product type numbers from ProductData and saves using CFSProductTypeForm"""

    if isinstance(product, CFSProduct):
        product_type_form_cls = CFSProductTypeForm
    else:
        product_type_form_cls = CFSProductTypeTemplateForm

    product_type_numbers = product_data.product_type_numbers
    existing_product_type_numbers = product.product_type_numbers.values_list(
        "product_type_number", flat=True
    )
    for product_type_number in product_type_numbers:
        if product_type_number not in existing_product_type_numbers:
            product_type_form = product_type_form_cls(
                data={"product_type_number": product_type_number}, product=product
            )

            if not product_type_form.is_valid():
                errors = product_type_form.errors
                msg = f"product type number '{product_type_number}'"

                if "product_type_number" in errors:
                    msg += f" - {errors['product_type_number'][0]}"

                raise CustomError(f"Product '{product.product_name} has error with {msg}")

            product_type_form.save()


def _process_ingredients(product: CFSProduct, product_data: ProductData) -> None:
    """Gets ingredeient data from ProductData and saves using CFSActiveIngredientForm"""

    if isinstance(product, CFSProduct):
        ingredient_form_cls = CFSActiveIngredientForm
    else:
        ingredient_form_cls = CFSActiveIngredientTemplateForm

    ingredient_names = product_data.ingredient_names
    cas_numbers = product_data.cas_numbers

    # ingredient_names and cas_numbers are both appended to for each row of data
    # the index for each ingredient will be the same in each list
    ingredient_data = zip(ingredient_names, cas_numbers)

    for ingredient in ingredient_data:
        name, cas_number = ingredient

        ingredient_form = ingredient_form_cls(
            data={"name": name, "cas_number": cas_number}, product=product
        )

        if not ingredient_form.is_valid():
            errors = ingredient_form.errors

            if "name" in errors:
                msg = f"active ingredient name '{name}' - {errors['name'][0]}"
            elif "cas_number" in errors:
                msg = f"CAS number '{cas_number}' - {errors['cas_number'][0]}"
            else:
                msg = f"active ingredient name '{name}' CAS number '{cas_number}'"

            raise CustomError(f"Product '{product.product_name}' has error with {msg}")

        ingredient_form.save()


def copy_export_application_to_export_application(
    existing_app: (
        CertificateOfFreeSaleApplication
        | CertificateOfManufactureApplication
        | CertificateOfGoodManufacturingPracticeApplication
    ),
    exporter: Exporter,
    exporter_office: Office,
    agent: Exporter | None,
    agent_office: Office | None,
    created_by: User,
) -> (
    CertificateOfFreeSaleApplication
    | CertificateOfManufactureApplication
    | CertificateOfGoodManufacturingPracticeApplication
):
    """Create a new export application from an existing one."""

    match existing_app:
        case CertificateOfFreeSaleApplication():
            model_cls = CertificateOfFreeSaleApplication
            copy_cls: type[ApplicationCopy] = CFSApplicationCopy

        case CertificateOfManufactureApplication():
            model_cls = CertificateOfManufactureApplication
            copy_cls = COMApplicationCopy

        case CertificateOfGoodManufacturingPracticeApplication():
            model_cls = CertificateOfGoodManufacturingPracticeApplication
            copy_cls = GMPApplicationCopy

        case _:
            raise ValueError(f"Can't create application for: {existing_app}")

    # Create the new model with basic application data.
    new_app = model_cls.objects.create(
        exporter=exporter,
        exporter_office=exporter_office,
        agent=agent,
        agent_office=agent_office,
        process_type=existing_app.process_type,
        application_type=existing_app.application_type,
        created_by=created_by,
        last_updated_by=created_by,
    )

    copier = copy_cls(existing_app, new_app, created_by)
    copier.copy_application_to_application()

    return new_app


def copy_export_application_to_template(
    cat: CertificateApplicationTemplate, application: ExportApplication
) -> None:
    """Update a certificate application template with export application data.

    :param cat: Certificate Application Template.
    :param application: Existing Export Application.
    """

    template, copy_cls = _get_template_and_copy_cls(cat)
    copier = copy_cls(application.get_specific_model(), template, template.template.owner)
    copier.copy_application_to_template()


def copy_template_to_export_application(
    application: (
        CertificateOfFreeSaleApplication
        | CertificateOfManufactureApplication
        | CertificateOfGoodManufacturingPracticeApplication
    ),
    cat: CertificateApplicationTemplate,
    user: User,
) -> None:
    """Update an export application with certificate application template data.

    :param application: Export Application.
    :param cat: Certificate Application Template.
    :param user: User who is creating the template.
    """

    template, copy_cls = _get_template_and_copy_cls(cat)
    copier = copy_cls(template, application, user)
    copier.copy_template_to_application()


def _get_template_and_copy_cls(
    cat: CertificateApplicationTemplate,
) -> tuple[
    CertificateOfFreeSaleApplicationTemplate
    | CertificateOfManufactureApplicationTemplate
    | CertificateOfGoodManufacturingPracticeApplicationTemplate,
    type[ApplicationCopy],
]:
    match cat.application_type:
        case ExportApplicationType.Types.FREE_SALE:
            template = cat.cfs_template
            copy_cls: type[ApplicationCopy] = CFSApplicationCopy

        case ExportApplicationType.Types.MANUFACTURE:
            template = cat.com_template
            copy_cls = COMApplicationCopy

        case ExportApplicationType.Types.GMP:
            template = cat.gmp_template
            copy_cls = GMPApplicationCopy

        case _:
            raise ValueError(f"Unknown application type {cat.application_type}")

    return template, copy_cls
